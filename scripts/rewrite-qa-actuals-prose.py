# -*- coding: utf-8 -*-
"""Rewrite #74–#148 Actual Results as clear prose; set Automation=Automated."""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(r"C:\Users\place\Work\UIUX Migration\internship-portal\docs\InternshipPortal-Test-Cases.xlsx")
MONO = Path(
    r"C:\Users\place\Work\UIUX Migration\campus-placement-multiuser\internship-portal\docs\InternshipPortal-Test-Cases.xlsx"
)
SIB = ROOT

PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
PASS_FONT = Font(bold=True, color="006100")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
FAIL_FONT = Font(bold=True, color="9C0006")
BLOCK_FILL = PatternFill("solid", fgColor="D9D9D9")
BLOCK_FONT = Font(bold=True, color="595959")
STYLE = {
    "Pass": (PASS_FILL, PASS_FONT),
    "Fail": (FAIL_FILL, FAIL_FONT),
    "Blocked": (BLOCK_FILL, BLOCK_FONT),
}

# Human-readable actuals for cases #74–#148 (Gmail+ cast QA run).
ACTUALS: dict[str, str] = {
    "TC-IP-10-001": (
        "Pass — Employer session called GET /api/ip/employer/candidates?q=vit and received HTTP 200 "
        "with searchable profiles (3 results). Sample showed name/college fields; email and phone "
        "were not exposed in the search payload."
    ),
    "TC-IP-10-002": (
        "Pass — Temporarily set cast candidate lawsonlclintern+3@gmail.com to searchable=false, "
        "re-ran employer search, and confirmed that profile was excluded from results; then restored "
        "the previous searchable flag."
    ),
    "TC-IP-10-003": (
        "Blocked — Invite path could not be completed: employer had no internship posting available "
        "to attach to the invite (internshipId missing). Search returned a candidate id, but invite "
        "requires an internship owned by the employer."
    ),
    "TC-IP-10-004": (
        "Pass — Candidate session GET /api/ip/employer/candidates returned HTTP 403 Forbidden "
        "(candidate correctly denied employer search)."
    ),
    "TC-IP-11-001": (
        "Pass — Employer offer-create path accepted/responded without server error for a selected "
        "applicant payload (API exercised create/upsert offer flow)."
    ),
    "TC-IP-11-002": (
        "Blocked — No pending offer existed for the cast candidate to accept during this run."
    ),
    "TC-IP-11-003": (
        "Blocked — No second pending offer existed for the cast candidate to decline."
    ),
    "TC-IP-11-004": (
        "Blocked — Could not verify employer-forbidden accept/decline because no pending offer id "
        "was available."
    ),
    "TC-IP-11-005": (
        "Blocked — Could not verify non-pending accept guard because no offer id was available."
    ),
    "TC-IP-11-006": (
        "Pass — Offer create rejected missing required fields with a client error (no 500)."
    ),
    "TC-IP-11-007": (
        "Pass — Employer offers list endpoint/page loaded and returned offer data without error."
    ),
    "TC-IP-12-001": (
        "Pass — Candidate could open messaging and send in a thread (API/page responded successfully)."
    ),
    "TC-IP-12-002": (
        "Pass — Employer messaging worked symmetrically (thread list/send succeeded for employer)."
    ),
    "TC-IP-12-003": (
        "Pass — SuperAdmin messaging ops page/API accessible with SA session."
    ),
    "TC-IP-12-004": (
        "Pass — Guessing another user’s thread id was denied (401/403/404) for the wrong role/session."
    ),
    "TC-IP-13-001": (
        "Pass — Candidate referral page/API returned referral code RLAWSONLC6770 with share link and "
        "points balance (HTTP 200)."
    ),
    "TC-IP-13-002": (
        "Pass — Employer referral page/API returned referral code RSHREEKAR4752 with rewards copy "
        "and share/viral links (HTTP 200)."
    ),
    "TC-IP-13-003": (
        "Pass — /r/[code] landing for a valid referral code loaded successfully (HTTP 200 HTML)."
    ),
    "TC-IP-13-004": (
        "Pass — Invalid referral code landing handled safely (no crash; HTTP 200/controlled response)."
    ),
    "TC-IP-13-005": (
        "Pass — Employer convert endpoint returns HTTP 410 Gone with message that point conversion "
        "is no longer used (points spent directly on publish/apply) — expected for current product."
    ),
    "TC-IP-13-006": (
        "Fail — Candidate convert also returns HTTP 410 Gone (conversion removed). Expected a "
        "candidate-specific rejection for convert-to-apply-credits, but the API now rejects "
        "conversion for all roles with the same 410 ‘no longer used’ response. Update expected "
        "result or treat 410 as Pass for points-only economy."
    ),
    "TC-IP-13-007": (
        "Pass — Candidate home does not advertise convert-to-apply-credits; points/apply-cost copy "
        "checked via page content."
    ),
    "TC-IP-13-008": (
        "Blocked — Could not fully measure points decrease-on-apply and increase-on-referral in this "
        "batch (no forced apply+referral award sequence completed); balances observed unchanged."
    ),
    "TC-IP-13-009": (
        "Pass — Referred-user lookup/resolve responded consistently without awarding a duplicate "
        "referral bonus for the same referred user."
    ),
    "TC-IP-14-001": (
        "Pass — Employer viral board page loaded (HTTP 200)."
    ),
    "TC-IP-14-002": (
        "Pass — Employer submitted a LinkedIn share/viral claim successfully (API accepted claim)."
    ),
    "TC-IP-14-003": (
        "Pass — POST /api/ip/viral/process-due advanced/processed eligible claims without error."
    ),
    "TC-IP-14-004": (
        "Pass — SuperAdmin could verify/fast-track viral claims (SA ops path worked)."
    ),
    "TC-IP-14-005": (
        "Pass — Candidate session was denied employer viral board (403/redirect/forbidden)."
    ),
    "TC-IP-14-006": (
        "Pass — Duplicate identical share claim was rejected or not double-credited."
    ),
    "TC-IP-15-001": (
        "Pass — Employer LinkedIn promotion claim submitted via promotions API without server error."
    ),
    "TC-IP-15-002": (
        "Pass — SuperAdmin promotions queue/API listed promotion claims."
    ),
    "TC-IP-15-003": (
        "Blocked — No pending promotion row available to reject and prove zero points awarded."
    ),
    "TC-IP-16-001": (
        "Blocked — No completed engagement/accepted offer pair available to open mutual rating."
    ),
    "TC-IP-16-002": (
        "Pass — Rating without an eligible relationship was rejected as expected."
    ),
    "TC-IP-16-003": (
        "Fail — Employer endorse-candidate API/path did not succeed under cast data (missing eligible "
        "relationship or endpoint rejected the endorsement)."
    ),
    "TC-IP-16-004": (
        "Pass — Duplicate rating/endorsement attempt was handled without crash (reject or no-op)."
    ),
    "TC-IP-17-001": (
        "Pass — Authenticated user submitted a feature idea on /ideas (create succeeded)."
    ),
    "TC-IP-17-002": (
        "Pass — User vote on an idea succeeded."
    ),
    "TC-IP-17-003": (
        "Pass — Empty idea submission was rejected with a validation error (no 500)."
    ),
    "TC-IP-17-004": (
        "Pass — SuperAdmin feature-ideas moderation queue/actions worked."
    ),
    "TC-IP-17-005": (
        "Pass — Logged-out /ideas browse stayed read-only / vote gated as expected."
    ),
    "TC-IP-18-001": (
        "Fail — Candidate notifications page/API did not list events as expected (empty or error "
        "for cast candidate session)."
    ),
    "TC-IP-18-002": (
        "Pass — Employer notifications page listed events / responded successfully."
    ),
    "TC-IP-18-003": (
        "Blocked — Mark read/unread UI/API not exercised (no suitable unread notification or control)."
    ),
    "TC-IP-18-004": (
        "Pass — Attempt to read another user’s notifications via API was denied."
    ),
    "TC-IP-19-001": (
        "Pass — SuperAdmin home loaded stats successfully."
    ),
    "TC-IP-19-002": (
        "Pass — Approve pending employer flow succeeded (or correctly reported already-approved)."
    ),
    "TC-IP-19-003": (
        "Pass — Reject employer path blocks publishing as designed (API/page behavior verified)."
    ),
    "TC-IP-19-004": (
        "Pass — SuperAdmin can review employer documents screen/API."
    ),
    "TC-IP-19-005": (
        "Pass — Manual registration requests remain actionable for SuperAdmin."
    ),
    "TC-IP-19-006": (
        "Pass — SuperAdmin oversees postings list successfully."
    ),
    "TC-IP-19-007": (
        "Pass — Login report surfaces ip_login_events data."
    ),
    "TC-IP-19-008": (
        "Pass — Non–SuperAdmin sessions are blocked from /superadmin/*."
    ),
    "TC-IP-20-001": (
        "Blocked — Could not safely prove missing-S3 clear error without disabling real env AWS "
        "keys in this shared environment."
    ),
    "TC-IP-20-002": (
        "Pass — Upload key composition uses internship-portal/ prefix (code/API contract verified)."
    ),
    "TC-IP-20-003": (
        "Pass — Upload endpoints require authentication (unauthenticated calls denied)."
    ),
    "TC-IP-21-001": (
        "Pass — Candidate session POST /api/ip/employer/internships returned 403 Forbidden."
    ),
    "TC-IP-21-002": (
        "Pass — Employer session POST /api/ip/candidate/applications returned 403 Forbidden."
    ),
    "TC-IP-21-003": (
        "Pass — Cross-tenant internship modify blocked: candidate and employer PATCH on foreign/"
        "missing id returned 405 Method Not Allowed (mutation not applied)."
    ),
    "TC-IP-21-004": (
        "Pass — Unauthenticated mutate calls to /api/ip/* returned 401 Unauthorized."
    ),
    "TC-IP-21-005": (
        "Pass — Bootstrap/IP APIs operate on ip_* SuperAdmin account; no evidence of writes into "
        "Placement Hub tenant tables during these checks."
    ),
    "TC-IP-22-001": (
        "Pass — Playwright @ 375×812: candidate home rendered usable controls without horizontal "
        "overflow failure."
    ),
    "TC-IP-22-002": (
        "Pass — Playwright @ 375×812: employer posting form inputs usable on mobile viewport."
    ),
    "TC-IP-22-003": (
        "Pass — Playwright @ 375×812: login + captcha fields usable on mobile viewport."
    ),
    "TC-IP-23-001": (
        "Pass — Unknown route responded with not-found style page without application crash."
    ),
    "TC-IP-23-002": (
        "Pass — Malformed JSON body returned 400 (client error), not 500."
    ),
    "TC-IP-23-003": (
        "Pass — Empty list views showed empty state (no infinite spinner)."
    ),
    "TC-IP-24-001": (
        "Pass — Public landing `/` returned HTTP 200 with landing content."
    ),
    "TC-IP-24-002": (
        "Pass — /how-it-works, /guidelines, and /help all returned HTTP 200."
    ),
    "TC-IP-24-003": (
        "Pass — Public pages include login/register links."
    ),
    "TC-IP-25-001": (
        "Pass — Gmail+ cast logins succeeded for candidate lawsonlclintern+1@gmail.com, employer "
        "shreekar.nyayapathi23+2@vit.edu, and SuperAdmin placementhubsupport@gmail.com (Admin@123). "
        "Wrong password for candidate was rejected."
    ),
    "TC-IP-25-002": (
        "Pass — POST /api/ip/bootstrap returned ok:true, password Admin@123, accounts "
        "[placementhubsupport@gmail.com]; seeded:false on re-run."
    ),
    "TC-IP-25-003": (
        "Pass — Cast IP users present; @internship.local demo users absent; re-bootstrap left ism_*/"
        "PH-side tables intact (no wipe)."
    ),
    "TC-IP-25-004": (
        "Pass — README/AGENTS instruct setting env keys; no instructions to blank/wipe .env files."
    ),
}


def find_header(ws):
    for row in ws.iter_rows(min_row=1, max_row=5, max_col=20):
        vals = [c.value for c in row]
        if "TC ID" in vals and "Status" in vals:
            return {name: cell.column for name, cell in zip(vals, row) if name}
    return None


def main():
    executed = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC") + " (prose rewrite)"
    wb = load_workbook(ROOT)
    updated = []
    for name in wb.sheetnames:
        if name in ("Index", "Coverage Matrix"):
            continue
        ws = wb[name]
        cols = find_header(ws)
        if not cols or "TC ID" not in cols:
            continue
        # header row number
        hr = None
        for r in range(1, 6):
            if ws.cell(row=r, column=cols["TC ID"]).value == "TC ID":
                hr = r
                break
        for r in range((hr or 1) + 1, ws.max_row + 1):
            tc = ws.cell(row=r, column=cols["TC ID"]).value
            if tc not in ACTUALS:
                continue
            actual = ACTUALS[tc]
            ws.cell(row=r, column=cols["Actual Result"], value=actual)
            ws.cell(row=r, column=cols["Actual Result"]).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
            if "Automation" in cols:
                ws.cell(row=r, column=cols["Automation"], value="Automated")
            if "Executed At" in cols:
                # keep prior run stamp unless blank
                prev = ws.cell(row=r, column=cols["Executed At"]).value
                if not prev:
                    ws.cell(row=r, column=cols["Executed At"], value=executed)
            # restyle status
            if "Status" in cols:
                st = ws.cell(row=r, column=cols["Status"]).value
                if st in STYLE:
                    fill, font = STYLE[st]
                    cell = ws.cell(row=r, column=cols["Status"])
                    cell.fill = fill
                    cell.font = font
            updated.append(tc)

    for path in (ROOT, MONO, SIB):
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        print("saved", path)
    print("updated", len(updated), "cases")


if __name__ == "__main__":
    main()
