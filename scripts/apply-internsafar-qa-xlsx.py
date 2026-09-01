# -*- coding: utf-8 -*-
"""Apply qa-results.json onto InternSafar-Test-Cases.xlsx (Legacy ID + TC ID unified)."""
from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "test-cases" / "InternSafar-Test-Cases.xlsx"
RESULTS = ROOT / "test-cases" / "qa-results.json"
SKIP = frozenset({"Index", "Coverage", "Notes", "How to use"})
# Updated only via scripts/manual/* — never from run-internsafar-qa.mjs --apply
MANUAL_ONLY_TC_IDS = frozenset({"TC-IS-06-007"})

PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
PASS_FONT = Font(bold=True, color="006100", name="Calibri")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
FAIL_FONT = Font(bold=True, color="9C0006", name="Calibri")
BLOCK_FILL = PatternFill("solid", fgColor="D9D9D9")
BLOCK_FONT = Font(bold=True, color="595959", name="Calibri")
NOTRUN_FILL = PatternFill("solid", fgColor="FFF2CC")
STYLE = {
    "Pass": (PASS_FILL, PASS_FONT),
    "Fail": (FAIL_FILL, FAIL_FONT),
    "Blocked": (BLOCK_FILL, BLOCK_FONT),
    "Not Run": (NOTRUN_FILL, Font(name="Calibri")),
}


def find_header(ws):
    for r in range(1, 10):
        vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 20) + 1)]
        if "TC ID" in vals and "Status" in vals:
            cols = {str(v): i + 1 for i, v in enumerate(vals) if v}
            return r, cols
    return None, {}


def style_status(cell, status: str):
    fill, font = STYLE.get(status, (NOTRUN_FILL, Font(name="Calibri")))
    cell.value = status
    cell.fill = fill
    cell.font = font
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def main():
    payload = json.loads(RESULTS.read_text(encoding="utf-8"))
    legacy_cases = payload.get("cases") or {}
    extra = payload.get("byTcId") or {}
    unified = dict(legacy_cases)
    unified.update(extra)
    unified.update(payload.get("results") or {})
    executed = payload.get("executedAt") or datetime.now(timezone.utc).isoformat()

    wb = load_workbook(XLSX)
    updated = 0
    unmatched = []
    seen_legacy = set()

    for ws in wb.worksheets:
        if ws.title in SKIP:
            continue
        hr, cols = find_header(ws)
        if not hr:
            continue
        sc = cols.get("Status")
        ac = cols.get("Actual Result")
        ec = cols.get("Executed At")
        lc = cols.get("Legacy ID")
        tc = cols.get("TC ID")
        if not sc:
            continue
        for r in range(hr + 1, ws.max_row + 1):
            tc_id = ws.cell(r, tc).value if tc else None
            legacy = (ws.cell(r, lc).value if lc else None) or ""
            legacy = str(legacy).strip()
            if tc_id and str(tc_id) in MANUAL_ONLY_TC_IDS:
                continue
            rec = None
            if tc_id and str(tc_id) in unified:
                rec = unified[str(tc_id)]
            elif legacy and legacy in unified:
                rec = unified[legacy]
                seen_legacy.add(legacy)
            if not rec:
                continue
            status = rec.get("status") or "Not Run"
            actual = rec.get("actual")
            if actual is not None and not isinstance(actual, str):
                actual = json.dumps(actual)
            style_status(ws.cell(r, sc), status)
            if ac:
                ws.cell(r, ac).value = actual
                ws.cell(r, ac).alignment = Alignment(vertical="top", wrap_text=True)
            if ec:
                ws.cell(r, ec).value = executed
            updated += 1

    leftover = sorted(k for k in legacy_cases if k not in seen_legacy)
    wb.save(XLSX)
    print(
        json.dumps(
            {
                "updatedRows": updated,
                "resultCases": len(unified),
                "extraTcIds": len(extra),
                "unmatchedLegacyIds": leftover,
            }
        )
    )
    if leftover:
        unmatched.extend(leftover)
        print("Unmatched Legacy IDs (no row):", ", ".join(leftover[:40]))


if __name__ == "__main__":
    main()
