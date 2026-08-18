# -*- coding: utf-8 -*-
"""Apply Status cell colors + update sheet results for InternshipPortal-Test-Cases.xlsx."""
from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Working copy lives under internship-portal/docs (root xlsx archived to _archive-root-clutter)
ROOT = Path(r"C:\Users\place\Work\UIUX Migration\internship-portal\docs\InternshipPortal-Test-Cases.xlsx")
MONO = Path(
    r"C:\Users\place\Work\UIUX Migration\campus-placement-multiuser\internship-portal\docs\InternshipPortal-Test-Cases.xlsx"
)
SIB = ROOT

PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
PASS_FONT = Font(bold=True, color="006100")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
FAIL_FONT = Font(bold=True, color="9C0006")
BLOCK_FILL = PatternFill("solid", fgColor="D9D9D9")
BLOCK_FONT = Font(bold=True, color="595959")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")

STYLE = {
    "Pass": (PASS_FILL, PASS_FONT),
    "Fail": (FAIL_FILL, FAIL_FONT),
    "Blocked": (BLOCK_FILL, BLOCK_FONT),
}


def find_status_col(ws):
    for row in ws.iter_rows(min_row=1, max_row=5, max_col=20):
        for cell in row:
            if cell.value == "Status":
                return cell.row, cell.column
    return None, None


def apply_status_style(cell):
    val = (cell.value or "").strip() if isinstance(cell.value, str) else cell.value
    if val in STYLE:
        fill, font = STYLE[val]
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)


def ensure_conditional_formatting(ws, header_row, status_col, max_row):
    """Add CellIs rules for Pass/Fail/Blocked on Status column."""
    if max_row < header_row + 1:
        return
    col_letter = get_column_letter(status_col)
    end = max(max_row, header_row + 200)
    rng = f"{col_letter}{header_row + 1}:{col_letter}{end}"
    try:
        # Drop prior rules whose sqref overlaps this column
        for sqref in list(ws.conditional_formatting._cf_rules.keys()):
            if col_letter in str(sqref):
                del ws.conditional_formatting[sqref]
    except Exception:
        pass
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"Pass"'], fill=PASS_FILL, font=PASS_FONT),
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"Fail"'], fill=FAIL_FILL, font=FAIL_FONT),
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"Blocked"'], fill=BLOCK_FILL, font=BLOCK_FONT),
    )


def style_all_status_columns(wb):
    for name in wb.sheetnames:
        if name in ("Index", "Coverage Matrix"):
            continue
        ws = wb[name]
        hr, sc = find_status_col(ws)
        if not sc:
            continue
        # restyle header Status to match theme
        hcell = ws.cell(row=hr, column=sc)
        hcell.fill = HEADER_FILL
        hcell.font = HEADER_FONT
        max_r = ws.max_row or hr
        for r in range(hr + 1, max_r + 1):
            apply_status_style(ws.cell(row=r, column=sc))
        ensure_conditional_formatting(ws, hr, sc, max_r)


def update_results(wb, sheet_name: str, cases: dict, executed_at: str):
    ws = wb[sheet_name]
    hr, _ = find_status_col(ws)
    header = [c.value for c in ws[hr]]
    col = {name: i + 1 for i, name in enumerate(header) if name}
    updated = []
    for row in range(hr + 1, ws.max_row + 1):
        tc = ws.cell(row=row, column=col["TC ID"]).value
        if tc not in cases:
            continue
        r = cases[tc]
        ws.cell(row=row, column=col["Status"], value=r["status"])
        apply_status_style(ws.cell(row=row, column=col["Status"]))
        ws.cell(row=row, column=col["Actual Result"], value=r["actual"])
        ws.cell(row=row, column=col["Executed At"], value=executed_at)
        updated.append(tc)
    return updated


def save_all(wb):
    wb.save(ROOT)
    MONO.parent.mkdir(parents=True, exist_ok=True)
    wb.save(MONO)
    if SIB.parent.parent.exists():
        SIB.parent.mkdir(parents=True, exist_ok=True)
        wb.save(SIB)


def apply_payload(wb, data):
    """Apply single-sheet or multi-batch QA JSON into workbook."""
    updated = []
    if "batches" in data:
        executed = data.get("executedAt") or datetime.now(timezone.utc).isoformat()
        for batch in data["batches"]:
            sheet = batch["sheet"]
            cases = batch["cases"]
            when = batch.get("executedAt") or executed
            updated.extend(update_results(wb, sheet, cases, when))
    else:
        updated.extend(update_results(wb, data["sheet"], data["cases"], data["executedAt"]))
    return updated


if __name__ == "__main__":
    import sys

    # args: optional results json path; else just apply formatting
    wb = load_workbook(ROOT)
    style_all_status_columns(wb)
    if len(sys.argv) > 1:
        raw = Path(sys.argv[1]).read_text(encoding="utf-8-sig")
        start = raw.find("{")
        data = json.loads(raw[start:] if start >= 0 else raw)
        updated = apply_payload(wb, data)
        # re-ensure CF after updates
        style_all_status_columns(wb)
        print("updated", updated)
    save_all(wb)
    print("saved", ROOT)
