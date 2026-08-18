# -*- coding: utf-8 -*-
"""Generate InternshipPortal-Test-Cases.xlsx in PlacementHub-Test-Cases column style.
Authored for campus-placement-multiuser/internship-portal — not a copy of PH cases.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
# Working outputs under kept app trees (root xlsx lived here until archived)
OUT = Path(r"C:\Users\place\Work\UIUX Migration\internship-portal\docs\InternshipPortal-Test-Cases.xlsx")
OUT_SIBLING = OUT

OUT_COPY = Path(
    r"C:\Users\place\Work\UIUX Migration\campus-placement-multiuser\internship-portal\docs\InternshipPortal-Test-Cases.xlsx"
)

COLS = [
    "TC ID",
    "Module",
    "Feature",
    "Title",
    "Priority",
    "Type",
    "Role(s)",
    "Preconditions",
    "Test Steps",
    "Expected Result",
    "Test Data / Notes",
    "Automation",
    "Status",
    "Actual Result",
    "Executed At",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
INDEX_TITLE_FONT = Font(size=16, bold=True, color="1F4E79")
LINK_FONT = Font(color="0563C1", underline="single")

# Demo seeds from README / prior QA
DEMO = {
    "candidate": "candidate@internship.local / Admin@123",
    "employer": "employer@internship.local / Admin@123",
    "superadmin": "superadmin@internship.local / Admin@123",
    "yopmail": "internship-portal-test@yopmail.com (OUTBOUND_EMAIL_OVERRIDE)",
}


def tc(
    mid: str,
    n: int,
    module: str,
    feature: str,
    title: str,
    priority: str,
    typ: str,
    roles: str,
    pre: str,
    steps: str,
    expected: str,
    notes: str = "",
    automation: str = "Manual",
):
    return {
        "id": f"TC-IP-{mid}-{n:03d}",
        "module": module,
        "feature": feature,
        "title": title,
        "priority": priority,
        "type": typ,
        "roles": roles,
        "pre": pre,
        "steps": steps,
        "expected": expected,
        "notes": notes,
        "automation": automation,
    }


def build_cases() -> dict[str, list[dict]]:
    """module_key -> list of cases. Sheet titles derived separately."""
    m: dict[str, list[dict]] = {}

    # ---- 01 Auth & Access ----
    mod = "01 Auth & Access"
    mid = "01"
    cases = [
        tc(mid, 1, mod, "Login", "Valid candidate credentials land on candidate home", "P0", "Functional", "Candidate",
           "Seeded candidate account", "1. Open / (home sign-in — NOT /login UI)\n2. Enter candidate email/password\n3. Solve captcha\n4. Submit",
           "Redirect to /candidate; candidate nav visible; session established", DEMO["candidate"]),
        tc(mid, 2, mod, "Login", "Valid employer credentials land on employer home", "P0", "Functional", "Employer",
           "Approved employer account", "1. Open /\n2. Sign in as employer with captcha",
           "Redirect to /employer; employer menus visible", DEMO["employer"]),
        tc(mid, 3, mod, "Login", "Invalid password shows error and no session", "P0", "Negative", "Public",
           "Known candidate email", "1. Enter correct email + wrong password\n2. Valid captcha\n3. Submit",
           "Error shown; remain unauthenticated; stay on home sign-in", ""),
        tc(mid, 4, mod, "Captcha", "Missing/wrong captcha blocks login on home /", "P0", "Negative", "Public",
           "Valid credentials; captcha on / and /superadmin/login (enforced in NextAuth)",
           "1. Enter valid creds on /\n2. Leave captcha blank or wrong\n3. Submit",
           "Login rejected; captcha error or challenge refreshed", "Captcha tokens use ~ separator (Vercel-safe)"),
        tc(mid, 5, mod, "Captcha", "Captcha challenge refreshes without full page break", "P1", "Functional", "Public",
           "/ open with sign-in", "1. Load captcha\n2. Request refresh/new challenge\n3. Solve new value and login",
           "New challenge works; login succeeds with new answer", ""),
        tc(mid, 6, mod, "Login redirect", "/login redirects to home sign-in preserving query", "P0", "Functional", "Public",
           "None", "1. Open /login\n2. Open /login?email=x&next=/candidate",
           "Client replace to / (and /?email=…&next=…); no separate login form page", "src/app/login/page.js"),
        tc(mid, 7, mod, "Session", "Unauthenticated visit to /candidate hits auth gate", "P0", "Security", "Public",
           "No session cookie", "1. Open /candidate in private window", "Redirect/gate to sign-in on /", ""),
        tc(mid, 8, mod, "Session", "Candidate cannot open /employer home as candidate", "P0", "Security", "Candidate",
           "Candidate session", "1. Navigate to /employer", "Denied or redirected to candidate home", ""),
        tc(mid, 9, mod, "Session", "Employer cannot open /superadmin without SA role", "P0", "Security", "Employer",
           "Employer session", "1. Open /superadmin", "Denied / redirect to employer or SA login", ""),
        tc(mid, 10, mod, "SuperAdmin login", "SA login via /superadmin/login with captcha", "P0", "Functional", "SuperAdmin",
           "Seeded SA", "1. Open /superadmin/login\n2. Enter SA creds + captcha",
           "Land on /superadmin dashboard", DEMO["superadmin"]),
        tc(mid, 11, mod, "Logout", "Sign-out clears session and blocks role home", "P0", "Functional", "Candidate",
           "Logged-in candidate", "1. Sign out\n2. Open /candidate", "Session cleared; forced to sign-in", ""),
        tc(mid, 12, mod, "Role home", "/app hub routes signed-in user to role home", "P1", "Functional", "Candidate, Employer",
           "Active session", "1. Open /app while logged in", "Redirect to matching role dashboard", ""),
        tc(mid, 13, mod, "Login", "Unknown email rejected", "P1", "Negative", "Public",
           "None", "1. Unknown email + any password + captcha on /", "Error; no session", ""),
        tc(mid, 14, mod, "Bootstrap", "POST /api/ip/bootstrap seeds demo trio idempotently", "P1", "Functional", "Developer / QA",
           "Dev/allowed env", "1. POST /api/ip/bootstrap\n2. Login each returned account",
           "Returns ok + accounts list; password Admin@123; seeded:false on re-run", "Verified live 200 ok seeded:false"),
    ]
    m[mod] = cases

    # ---- 02 Registration ----
    mod = "02 Registration"
    mid = "02"
    cases = [
        tc(mid, 1, mod, "Candidate register", "Gmail candidate registration sends temp password email", "P0", "Functional", "Public",
           "Mail override or SMTP configured", "1. Open /register/candidate\n2. Complete form with @gmail.com\n3. Submit",
           "Account created; temp password email sent (to override inbox if set)", DEMO["yopmail"]),
        tc(mid, 2, mod, "Candidate register", "Non-Gmail candidate signup blocked or refused", "P0", "Negative", "Public",
           "/register/candidate", "1. Attempt signup with non-gmail address", "Validation error; no account", "Product: Gmail-only candidates"),
        tc(mid, 3, mod, "Candidate register", "Duplicate email rejected", "P0", "Negative", "Public",
           "Existing candidate@internship.local", "1. Register again with same email", "Clear duplicate error", ""),
        tc(mid, 4, mod, "Employer register", "Employer domain-match path creates pending/approved employer", "P0", "Functional", "Public",
           "Valid company email domain path enabled", "1. Open /register/employer\n2. Complete required fields\n3. Submit",
           "Employer user created; status pending or approved per rules; email with creds if applicable", ""),
        tc(mid, 5, mod, "Employer register", "Employer manual request reaches SuperAdmin queue", "P0", "Functional", "Public, SuperAdmin",
           "Manual request path", "1. Submit employer request without auto-approve domain\n2. Open SuperAdmin requests/approvals",
           "Request visible for SA action", ""),
        tc(mid, 6, mod, "Google button", "Google Sign-In UI starts credentials-style registration flow (not live OAuth)", "P1", "Functional", "Public",
           "/register or /login", "1. Use Google-looking control\n2. Observe flow",
           "Flows through credentials/temp-password registration (fake OAuth UX per product)", "Do not expect real Google OAuth tokens"),
        tc(mid, 7, mod, "Validation", "Required fields empty block candidate register", "P1", "Negative", "Public",
           "/register/candidate", "1. Submit empty form", "Field validation errors; no API create", ""),
        tc(mid, 8, mod, "Register hub", "/register offers candidate vs employer paths", "P1", "UI", "Public",
           "None", "1. Open /register", "Clear links to /register/candidate and /register/employer", ""),
    ]
    m[mod] = cases

    # ---- 03 Password Reset & Email ----
    mod = "03 Password Reset & Email"
    mid = "03"
    cases = [
        tc(mid, 1, mod, "Forgot password", "Forgot password sends reset mail to override inbox", "P0", "Functional", "Public",
           "Known user; OUTBOUND_EMAIL_OVERRIDE set", "1. Open /forgot-password\n2. Enter known email\n3. Submit\n4. Check YOPmail",
           "Success message; email received at override address", DEMO["yopmail"]),
        tc(mid, 2, mod, "Forgot password", "Unknown email does not reveal account existence harshly", "P1", "Security", "Public",
           "/forgot-password", "1. Submit unknown email", "Generic success or safe error; no user enumeration detail", ""),
        tc(mid, 3, mod, "Reset", "Valid reset token on /forgot-password?token= allows password change", "P0", "Functional", "Candidate",
           "Fresh reset token from email", "1. Open /forgot-password?token=…\n2. Set new password\n3. Login with new password",
           "Password updated; old password fails; new works (no separate /reset-password route)", ""),
        tc(mid, 4, mod, "Reset", "Expired/invalid reset token rejected", "P0", "Negative", "Public",
           "Tampered or old token", "1. Open /forgot-password?token=bad\n2. Attempt change", "Error; password unchanged", ""),
        tc(mid, 5, mod, "Mail fallback", "When override off and primary fails, IP_MAIL_TEST_FALLBACK used (if enabled)", "P2", "Edge", "Developer / QA",
           "Controlled mail failure scenario", "1. Trigger outbound mail without override\n2. Force failure path",
           "Fallback inbox receives mail OR clear log of failure when fallback disabled", "Set IP_MAIL_TEST_FALLBACK=0 to disable"),
        tc(mid, 6, mod, "Temp password", "Candidate can change temp password after first login", "P1", "Functional", "Candidate",
           "Fresh registered candidate with temp password", "1. Login with temp\n2. Change via change-password API/UI",
           "New password works thereafter", ""),
    ]
    m[mod] = cases

    # ---- 04 Candidate Home & Profile ----
    mod = "04 Candidate Home & Profile"
    mid = "04"
    cases = [
        tc(mid, 1, mod, "Home", "Candidate home loads without fatal errors", "P0", "Functional", "Candidate",
           DEMO["candidate"], "1. Sign in\n2. Open /candidate", "Dashboard widgets/stats load; no 500", ""),
        tc(mid, 2, mod, "Profile", "View and save candidate profile fields", "P0", "Functional", "Candidate",
           "Candidate session", "1. Open /candidate/profile\n2. Edit name/education/skills\n3. Save",
           "Success toast; values persist on reload", ""),
        tc(mid, 3, mod, "Photo", "Upload candidate photo to S3 under internship-portal prefix", "P0", "Functional", "Candidate",
           "AWS/S3 env configured", "1. Upload JPG/PNG photo\n2. Reload profile",
           "Photo displays; object stored under internship-portal/…", ""),
        tc(mid, 4, mod, "Photo", "Reject unsupported photo type/oversize", "P1", "Negative", "Candidate",
           "Profile upload UI", "1. Upload exe or huge file", "Rejected with clear error; prior photo kept", ""),
        tc(mid, 5, mod, "Searchable", "Opt-in searchable profile toggles visibility to employers", "P0", "Functional", "Candidate, Employer",
           "Both sessions", "1. Candidate enables searchable\n2. Employer opens candidate search",
           "Candidate appears when searchable; hidden when off", ""),
        tc(mid, 6, mod, "Export", "Candidate profile export downloads CSV (labeled Excel)", "P2", "Functional", "Candidate",
           "Profile page", "1. Trigger Export Excel (.csv) control\n2. Open downloaded file",
           "CSV downloads successfully (not .xlsx); columns match API export", "UI label Excel (.csv); /api/ip/candidate/export returns CSV"),
        tc(mid, 7, mod, "Validation", "Invalid profile fields rejected (e.g. empty required)", "P1", "Negative", "Candidate",
           "Profile edit", "1. Clear required fields\n2. Save", "Validation errors; no partial corrupt save", ""),
        tc(mid, 8, mod, "Reminder", "Profile reminder banner shows on incomplete profile milestones", "P1", "Functional", "Candidate",
           "Incomplete profile; login counts hit 1/3/7/14/30", "1. Open any candidate page under PortalShell",
           "ProfileReminderBanner appears when incomplete per profileReminder.js milestones", "Milestones [1,3,7,14,30]"),
        tc(mid, 9, mod, "Change password", "Authenticated change-password API updates credentials", "P1", "Functional", "Candidate",
           "Candidate session; known current password", "1. Call /api/ip/auth/change-password (or UI if present)\n2. Login with new password",
           "Old password fails; new works", ""),
    ]
    m[mod] = cases

    # ---- 05 Browse Save Apply ----
    mod = "05 Browse Save Apply"
    mid = "05"
    cases = [
        tc(mid, 1, mod, "Browse", "Candidate internship list loads published/approved postings", "P0", "Functional", "Candidate",
           "At least one live posting", "1. Open /candidate/internships", "List shows eligible postings; draft/unapproved absent", ""),
        tc(mid, 2, mod, "Detail", "Internship detail page shows description and apply CTA", "P0", "Functional", "Candidate",
           "Known posting id", "1. Open /candidate/internships/[id]", "Detail content loads; apply/save available when eligible", ""),
        tc(mid, 3, mod, "Save", "Candidate can save/unsave internship", "P1", "Functional", "Candidate",
           "Detail page", "1. Save\n2. Confirm saved state\n3. Unsave", "Persists across reload", ""),
        tc(mid, 4, mod, "Apply", "Apply always spends POINTS_PER_APPLICATION points (5)", "P0", "Functional", "Candidate",
           "Candidate with points >= 5; profile_complete true", "1. Apply to eligible posting\n2. Confirm message + points balance",
           "Application created; points decreased by 5; success message shows spent points",
           "pointsEconomy POINTS_PER_APPLICATION=ceil(25/5)=5; api/ip/candidate/applications"),
        tc(mid, 5, mod, "Apply", "application_allowance is awarded/shown but NOT consumed on apply", "P0", "Regression", "Candidate",
           "Candidate with application_allowance > 0 and enough points", "1. Note allowance on detail UI\n2. Apply\n3. Re-check allowance and points",
           "Points decrease by 5; application_allowance unchanged (column is reward display only today)",
           "CRITICAL: suite must NOT claim allowance-before-points — apply path ignores allowance"),
        tc(mid, 6, mod, "Apply", "Apply blocked when points < POINTS_PER_APPLICATION", "P0", "Negative", "Candidate",
           "points < 5; profile complete", "1. Attempt apply", "Clear need-N-points error; no application row", ""),
        tc(mid, 7, mod, "Apply", "Duplicate apply to same internship rejected", "P0", "Negative", "Candidate",
           "Already applied", "1. Apply again", "Rejected; single application retained", ""),
        tc(mid, 8, mod, "Visibility", "Unapproved/pending employer posting not visible to candidates", "P0", "Security", "Candidate",
           "Pending posting exists", "1. Browse list and try direct id if known", "Not listed; direct access denied/empty", ""),
        tc(mid, 9, mod, "Filters", "Search/filters narrow internship list if UI provides", "P1", "Functional", "Candidate",
           "Multiple postings", "1. Apply keyword/filter", "Results match criteria", ""),
        tc(mid, 10, mod, "Profile gate", "Apply blocked when candidate profile_complete is false", "P0", "Negative", "Candidate",
           "Incomplete candidate profile", "1. Attempt apply", "Blocked until profile complete", "applications/route.js profileGate"),
    ]
    m[mod] = cases

    # ---- 06 Candidate Applications ----
    mod = "06 Candidate Applications"
    mid = "06"
    cases = [
        tc(mid, 1, mod, "List", "My applications lists prior applies with status", "P0", "Functional", "Candidate",
           ">=1 application", "1. Open /candidate/applications", "Rows show posting + status", ""),
        tc(mid, 2, mod, "Withdraw", "Withdraw control shown only for applied status in UI", "P1", "Functional", "Candidate",
           "Application status=applied", "1. Open /candidate/applications\n2. Withdraw\n3. Reload",
           "Status becomes withdrawn; employer sees withdrawn", "UI gates button to applied"),
        tc(mid, 3, mod, "Withdraw", "API DELETE/withdraw on owned application succeeds without status gate", "P1", "Edge", "Candidate, Developer / QA",
           "Non-applied owned application id (if obtainable)", "1. Call withdraw API for owned id",
           "API currently allows withdraw without terminal-status block (UI may still hide button)",
           "Do not claim product forbids terminal withdraw at API — only UI hides"),
        tc(mid, 4, mod, "Sync", "Status changes by employer appear on candidate applications", "P0", "Integration", "Candidate, Employer",
           "Employer updates pipeline status", "1. Employer sets shortlisted/interviewing/etc.\n2. Candidate refreshes",
           "Status matches employer pipeline values (applied, shortlisted, interviewing, rejected, hired, completed, offered, …)", ""),
        tc(mid, 5, mod, "Statuses", "Employer pipeline exposes applied→…→hired/completed transitions", "P0", "State", "Employer",
           "Applicant on posting", "1. Open /employer/internships/[id]\n2. Cycle statuses available in UI",
           "Status updates persist", "Includes offered in UI; offer accept may set hired"),
    ]
    m[mod] = cases

    # ---- 07 Employer Profile Ethics Docs ----
    mod = "07 Employer Profile Ethics Docs"
    mid = "07"
    cases = [
        tc(mid, 1, mod, "Home", "Employer home loads", "P0", "Functional", "Employer",
           DEMO["employer"], "1. Sign in\n2. Open /employer", "Dashboard loads; points/credits shown if present", ""),
        tc(mid, 2, mod, "Profile", "Update company profile and save", "P0", "Functional", "Employer",
           "Employer session", "1. Open /employer/profile\n2. Edit fields\n3. Save", "Persists on reload", ""),
        tc(mid, 3, mod, "Ethics", "Ethics checklist gates profile completion / posting readiness", "P0", "Functional", "Employer",
           "Ethics not fully acked", "1. Attempt publish/posting while incomplete\n2. Complete ethics acks",
           "Blocked until complete; allowed after", "Ethics as profile-completion gate"),
        tc(mid, 4, mod, "Logo", "Upload employer logo to S3", "P0", "Functional", "Employer",
           "S3 configured", "1. Upload logo\n2. Reload", "Logo shown; stored under internship-portal/…", ""),
        tc(mid, 5, mod, "Documents", "Upload employer verification documents", "P0", "Functional", "Employer",
           "Docs UI", "1. Upload PDF docs\n2. SuperAdmin documents queue", "Docs appear for SA review", ""),
        tc(mid, 6, mod, "Documents", "Reject bad document types", "P1", "Negative", "Employer",
           "Upload UI", "1. Upload unsupported type", "Rejected", ""),
        tc(mid, 7, mod, "Approval", "Unapproved employer cannot create ANY internship via API (including draft)", "P0", "Security", "Employer",
           "Employer approval_status != approved", "1. POST create internship (draft or publish)\n2. Observe UI copy saying drafts OK while pending",
           "API returns 403 until approved; dashboard copy that allows drafts while pending is OVERSTATED vs API",
           "employer/internships/route.js approval_status gate"),
    ]
    m[mod] = cases

    # ---- 08 Employer Postings & Credits ----
    mod = "08 Employer Postings & Credits"
    mid = "08"
    cases = [
        tc(mid, 1, mod, "Create", "Approved employer with profile_complete creates internship (draft/publish)", "P0", "Functional", "Employer",
           "approval_status=approved AND profile_complete", "1. Open /employer/internships/new\n2. Fill required\n3. Save",
           "Appears in /employer/internships; visibility to candidates depends on publish rules", ""),
        tc(mid, 2, mod, "Publish", "Publish consumes free post credit when available", "P0", "Functional", "Employer",
           "freePostCredits >= 1; approved+complete", "1. Publish posting\n2. Check credits", "Credit decremented; posting live/pending per rules", ""),
        tc(mid, 3, mod, "Publish", "Publish blocked with 0 credits and insufficient points to convert", "P0", "Negative", "Employer",
           "0 freePostCredits; points < 50", "1. Attempt publish",
           "Blocked with convert/credits message", "POINTS_PER_FREE_POST_CREDIT=50"),
        tc(mid, 4, mod, "Convert", "Employer converts points to free post credits (50→1)", "P0", "Functional", "Employer",
           "points >= 50", "1. Use points convert UI/API\n2. Observe credits + points",
           "Points −50 (or multiple); freePostCredits +1 per conversion", "/api/ip/points/convert employer-only"),
        tc(mid, 5, mod, "Edit", "Employer edits existing internship via edit route", "P1", "Functional", "Employer",
           "Owned posting", "1. Open /employer/internships/[id]/edit\n2. Change title\n3. Save", "Updates reflected on detail", ""),
        tc(mid, 6, mod, "Validation", "Create rejects missing required fields / bad dates", "P0", "Negative", "Employer",
           "New form", "1. Submit empty or end-before-start", "Validation errors", ""),
        tc(mid, 7, mod, "List", "Employer internships list shows owned postings only", "P0", "Security", "Employer",
           "Two employers with postings", "1. List as employer A", "Only A postings listed", ""),
        tc(mid, 8, mod, "Analytics", "Employer analytics page loads live SQL aggregates", "P1", "Functional", "Employer",
           "Postings with applicants", "1. Open /employer/analytics", "Stats from /api/ip/employer/analytics; empty state ok; no 500", "Rule-of-thumb AI insight text is not ML"),
        tc(mid, 9, mod, "Profile gate", "Employer without profile_complete cannot create posting", "P0", "Negative", "Employer",
           "Approved but ethics/profile incomplete", "1. Attempt POST create", "403 until profile_complete (ethics included)", ""),
        tc(mid, 10, mod, "Completions", "Mark completion path via completions API/UI when eligible", "P1", "Functional", "Employer",
           "Hired/eligible applicant", "1. Mark completed\n2. Confirm status", "Completion recorded; may unlock ratings", "/api/ip/completions if exposed"),
    ]
    m[mod] = cases

    # ---- 09 Employer Applicants Pipeline ----
    mod = "09 Employer Applicants Pipeline"
    mid = "09"
    cases = [
        tc(mid, 1, mod, "Applicants", "Employer sees applicants on posting detail", "P0", "Functional", "Employer",
           "Posting with applies", "1. Open /employer/internships/[id]", "Applicant list with statuses", ""),
        tc(mid, 2, mod, "Pipeline", "Employer can progress applicant status (shortlist/select/reject)", "P0", "Functional", "Employer",
           "Applied candidate", "1. Change status\n2. Candidate refreshes applications", "Both sides reflect new status", ""),
        tc(mid, 3, mod, "Complete", "Mark hire/completion path when product supports", "P1", "Functional", "Employer",
           "Selected applicant", "1. Complete hire action", "Terminal status set; downstream offer/rating eligible if gated", ""),
        tc(mid, 4, mod, "Auth", "Employer cannot view another employer applicants API", "P0", "Security", "Employer",
           "Other employer posting id", "1. Call applicants API/UI for foreign id", "403/empty", ""),
    ]
    m[mod] = cases

    # ---- 10 Candidate Search & Invite ----
    mod = "10 Candidate Search & Invite"
    mid = "10"
    cases = [
        tc(mid, 1, mod, "Search", "Employer candidate search finds searchable profiles", "P0", "Functional", "Employer",
           "Searchable candidate exists", "1. Open /employer/candidates\n2. Search", "Matching candidates shown", ""),
        tc(mid, 2, mod, "Privacy", "Non-searchable candidates excluded", "P0", "Security", "Employer, Candidate",
           "Candidate opted out", "1. Search for that profile", "Not returned", ""),
        tc(mid, 3, mod, "Invite", "Employer can invite candidate to apply", "P0", "Functional", "Employer, Candidate",
           "Live posting + searchable candidate", "1. Invite to apply\n2. Candidate notifications/messages",
           "Invite delivered; candidate can open target posting", ""),
        tc(mid, 4, mod, "Auth", "Candidate cannot open employer candidate search", "P0", "Security", "Candidate",
           "Candidate session", "1. Open /employer/candidates", "Denied", ""),
    ]
    m[mod] = cases

    # ---- 11 Offers ----
    mod = "11 Offers"
    mid = "11"
    cases = [
        tc(mid, 1, mod, "Create", "Employer creates offer for selected applicant", "P0", "Functional", "Employer",
           "Eligible application", "1. Create offer from /employer/offers or posting flow\n2. Candidate opens /candidate/offers",
           "Offer pending; visible to candidate", ""),
        tc(mid, 2, mod, "Accept", "Candidate accepts pending offer", "P0", "State", "Candidate",
           "Pending offer", "1. Accept on /candidate/offers",
           "Status accepted; may set application hired; employer sees accepted", "Accept/decline is candidate-only API"),
        tc(mid, 3, mod, "Decline", "Candidate declines pending offer", "P0", "State", "Candidate",
           "Pending offer", "1. Decline", "Offer declined; application may become declined_offer", ""),
        tc(mid, 4, mod, "Negative", "Employer cannot accept/decline as candidate", "P0", "Security", "Employer",
           "Pending offer id", "1. Call offers/[id] respond as employer", "Denied — respond is candidate-only", ""),
        tc(mid, 5, mod, "Negative", "Cannot accept non-pending offer", "P0", "Negative", "Candidate",
           "Already accepted/declined", "1. Accept again", "Error/conflict", ""),
        tc(mid, 6, mod, "Validation", "Offer create rejects missing critical fields", "P1", "Negative", "Employer",
           "Offers UI", "1. Submit incomplete offer", "Validation errors", ""),
        tc(mid, 7, mod, "List", "Employer offers list shows sent offers; ratings/endorsements after accept", "P1", "Functional", "Employer",
           ">=1 offer", "1. Open /employer/offers", "Offers listed; post-accept rating/endorse controls when eligible", ""),
    ]
    m[mod] = cases

    # ---- 12 Messages ----
    mod = "12 Messages"
    mid = "12"
    cases = [
        tc(mid, 1, mod, "Thread", "Candidate can open messages and send in a thread", "P0", "Functional", "Candidate",
           "Existing thread or start path", "1. Open /candidate/messages\n2. Open thread\n3. Send message",
           "Message persists; other party can read", ""),
        tc(mid, 2, mod, "Thread", "Employer messaging works symmetrically", "P0", "Functional", "Employer",
           "Thread with candidate", "1. Open /employer/messages/[id]\n2. Reply", "Candidate sees reply", ""),
        tc(mid, 3, mod, "SA", "SuperAdmin can access messaging ops page", "P1", "Functional", "SuperAdmin",
           "SA session", "1. Open /superadmin/messages", "Page loads; moderation/ops possible per UI", ""),
        tc(mid, 4, mod, "Security", "User cannot open another user's thread by id guess", "P0", "Security", "Candidate",
           "Foreign thread id", "1. Open /candidate/messages/[foreign]", "404/403", ""),
    ]
    m[mod] = cases

    # ---- 13 Points Referrals Convert ----
    mod = "13 Points Referrals Convert"
    mid = "13"
    cases = [
        tc(mid, 1, mod, "Referral page", "Candidate referral page shows code and share link", "P0", "Functional", "Candidate",
           "Candidate session", "1. Open /candidate/referral", "Code + /r/{code} link visible", ""),
        tc(mid, 2, mod, "Referral page", "Employer referral page shows code and rewards copy", "P0", "Functional", "Employer",
           "Employer session", "1. Open /employer/referral", "Code shown; employer reward described",
           "Employer reward: points + freePostCredits"),
        tc(mid, 3, mod, "Landing", "/r/[code] attribution landing works", "P0", "Functional", "Public",
           "Valid referral code", "1. Open /r/{code}\n2. Register accordingly",
           "Referral attributed; referrer +25 points; employer also +1 freePostCredit OR candidate +2 application_allowance (award only)",
           "Allowance is awarded to referred candidates' referrer path — not spent on apply"),
        tc(mid, 4, mod, "Landing", "Invalid referral code handled safely", "P1", "Negative", "Public",
           "None", "1. Open /r/not-a-real-code", "Friendly error or register without crash", ""),
        tc(mid, 5, mod, "Convert", "Only employers convert points to posting credits", "P0", "Functional", "Employer",
           "Enough points", "1. Convert via UI\n2. Check ledger", "Credits increase; points decrease by 50 each", ""),
        tc(mid, 6, mod, "Convert", "Candidate convert API is rejected (points spent on apply only)", "P0", "Negative", "Candidate",
           "Candidate with points", "1. POST /api/ip/points/convert as candidate",
           "Denied; no free-post conversion for candidates", "candidates blocked in convert/route.js"),
        tc(mid, 7, mod, "UI copy", "Candidate home must not imply convert-to-apply-credits", "P1", "Regression", "Candidate",
           "Candidate home", "1. Read points/cost cards on /candidate\n2. Compare to ReferralCard copy",
           "Cost-per-apply shows 5 pts; no candidate convert-to-apps step", "If UI says convert wrongly, file product bug — economy is apply-spend"),
        tc(mid, 8, mod, "Ledger", "Points decrease after apply and increase after referral award", "P1", "Functional", "Candidate",
           "Prior referral + apply", "1. Inspect points on home/profile", "Balances match actions", ""),
        tc(mid, 9, mod, "Double claim", "Same referred user cannot award referrer twice", "P0", "Negative", "Public",
           "Already attributed user", "1. Reuse code path again", "No double reward", ""),
    ]
    m[mod] = cases

    # ---- 14 Viral Board ----
    mod = "14 Viral Board"
    mid = "14"
    cases = [
        tc(mid, 1, mod, "Board", "Employer viral board page loads", "P0", "Functional", "Employer",
           "Approved employer", "1. Open /employer/viral", "Board UI loads; share/claim controls visible", ""),
        tc(mid, 2, mod, "Share claim", "Employer can submit LinkedIn share claim for points", "P0", "Functional", "Employer",
           "Viral UI", "1. Submit share/claim per UI", "Claim recorded as pending/verifying", ""),
        tc(mid, 3, mod, "Verify delay", "POST /api/ip/viral/process-due advances eligible claims", "P1", "Integration", "SuperAdmin, System",
           "Pending viral claim past due window", "1. Create claim\n2. Invoke process-due (SA/cron)\n3. Refresh employer viral",
           "Claim moves verified/rejected; rewards only when verified", "/api/ip/viral/process-due"),
        tc(mid, 4, mod, "SA fast-track", "SuperAdmin can verify/fast-track viral claims", "P0", "Functional", "SuperAdmin",
           "Pending claim", "1. Open /superadmin/viral\n2. Approve/verify", "Employer rewarded; claim closed", ""),
        tc(mid, 5, mod, "Auth", "Candidate cannot access employer viral board", "P0", "Security", "Candidate",
           "Candidate session", "1. Open /employer/viral", "Denied", ""),
        tc(mid, 6, mod, "Fraud", "Duplicate identical share claim rejected or queued once", "P1", "Negative", "Employer",
           "Existing identical claim", "1. Resubmit same claim", "No duplicate rewards", ""),
    ]
    m[mod] = cases

    # ---- 15 LinkedIn Promotions ----
    mod = "15 LinkedIn Promotions"
    mid = "15"
    cases = [
        tc(mid, 1, mod, "Claim", "Employer can claim LinkedIn promotion for points/credits", "P0", "Functional", "Employer",
           "Promo UI available", "1. Submit promo claim", "Pending review; LINKEDIN_PROMO_POINTS/CREDITS on approve",
           "30 points / 1 credit on approval"),
        tc(mid, 2, mod, "SA review", "SuperAdmin promotions queue lists claims", "P0", "Functional", "SuperAdmin",
           "Pending promo", "1. Open /superadmin/promotions\n2. Approve/reject", "Rewards applied only on approve", ""),
        tc(mid, 3, mod, "Reject", "Rejected promo grants no points", "P0", "Negative", "SuperAdmin, Employer",
           "Pending promo", "1. Reject", "Employer balance unchanged", ""),
    ]
    m[mod] = cases

    # ---- 16 Ratings & Endorsements ----
    mod = "16 Ratings & Endorsements"
    mid = "16"
    cases = [
        tc(mid, 1, mod, "Rate", "Mutual rating available after engagement/offer completion", "P0", "Functional", "Candidate, Employer",
           "Completed engagement per product rules", "1. Submit rating\n2. Other party views", "Rating stored; visible per rules", "/api/ip/ratings"),
        tc(mid, 2, mod, "Rate", "Cannot rate without eligible relationship", "P0", "Negative", "Candidate",
           "No prior engagement", "1. Attempt rate API/UI", "Denied", ""),
        tc(mid, 3, mod, "Endorse", "Employer can endorse candidate", "P1", "Functional", "Employer",
           "Eligible pair", "1. Create endorsement", "Endorsement on candidate profile; LinkedIn share optional", "/api/ip/endorsements"),
        tc(mid, 4, mod, "Duplicate", "Duplicate rating/endorsement handled", "P1", "Negative", "Employer",
           "Already rated", "1. Rate again", "Update-or-reject per product; no crash", ""),
    ]
    m[mod] = cases

    # ---- 17 Feature Ideas ----
    mod = "17 Feature Ideas"
    mid = "17"
    cases = [
        tc(mid, 1, mod, "Submit", "Authenticated user can submit feature idea on /ideas", "P0", "Functional", "Candidate, Employer",
           "Logged in", "1. Open /ideas\n2. Submit title+description", "Idea appears in list/pending", ""),
        tc(mid, 2, mod, "Vote", "User can vote on an idea", "P0", "Functional", "Candidate",
           "Existing idea", "1. Vote\n2. Reload", "Vote count increments; toggle off removes vote if supported", ""),
        tc(mid, 3, mod, "Validation", "Empty idea rejected", "P1", "Negative", "Candidate",
           "/ideas", "1. Submit blank", "Validation error", ""),
        tc(mid, 4, mod, "SA moderate", "SuperAdmin feature-ideas queue can moderate", "P0", "Functional", "SuperAdmin",
           "Pending ideas", "1. Open /superadmin/feature-ideas\n2. Approve/reject/status", "Status updates reflected on /ideas", ""),
        tc(mid, 5, mod, "Public", "Logged-out user browsing /ideas is read-only or gated for vote", "P1", "Security", "Public",
           "No session", "1. Open /ideas\n2. Attempt vote", "Vote requires auth", ""),
    ]
    m[mod] = cases

    # ---- 18 Notifications ----
    mod = "18 Notifications"
    mid = "18"
    cases = [
        tc(mid, 1, mod, "Inbox", "Candidate notifications page lists events", "P0", "Functional", "Candidate",
           "Triggered invite/offer/message", "1. Open /candidate/notifications", "Relevant notifications shown", ""),
        tc(mid, 2, mod, "Inbox", "Employer notifications page lists events", "P0", "Functional", "Employer",
           "New applicant event", "1. Open /employer/notifications", "Notification present", ""),
        tc(mid, 3, mod, "Read", "Mark read / unread behavior if UI supports", "P1", "Functional", "Candidate",
           "Unread items", "1. Open item\n2. Reload", "Unread count decreases", ""),
        tc(mid, 4, mod, "Isolation", "User cannot read another user's notifications via API", "P0", "Security", "Candidate",
           "Other user ids", "1. Call notifications API", "Only own rows", ""),
    ]
    m[mod] = cases

    # ---- 19 SuperAdmin Ops ----
    mod = "19 SuperAdmin Ops"
    mid = "19"
    cases = [
        tc(mid, 1, mod, "Dashboard", "SuperAdmin home loads stats", "P0", "Functional", "SuperAdmin",
           DEMO["superadmin"], "1. Open /superadmin", "Stats/overview load", ""),
        tc(mid, 2, mod, "Approvals", "Approve pending employer", "P0", "Functional", "SuperAdmin",
           "Pending employer", "1. Open /superadmin/approvals\n2. Approve", "Employer can publish thereafter", ""),
        tc(mid, 3, mod, "Approvals", "Reject employer blocks publishing", "P0", "Functional", "SuperAdmin, Employer",
           "Pending employer", "1. Reject\n2. Employer tries publish", "Still blocked", ""),
        tc(mid, 4, mod, "Documents", "Review employer documents", "P0", "Functional", "SuperAdmin",
           "Uploaded docs", "1. Open /superadmin/documents\n2. Approve/reject", "Status updated", ""),
        tc(mid, 5, mod, "Requests", "Manual registration requests actionable", "P0", "Functional", "SuperAdmin",
           "Pending request", "1. Open /superadmin/requests", "Can approve/deny", ""),
        tc(mid, 6, mod, "Postings", "SA can oversee postings list", "P1", "Functional", "SuperAdmin",
           "Live postings", "1. Open /superadmin/postings", "Postings listed; actions per UI", ""),
        tc(mid, 7, mod, "Login report", "Login report shows ip_login_events", "P1", "Functional", "SuperAdmin",
           "Prior logins", "1. Open /superadmin/login-report", "Events/table loads", ""),
        tc(mid, 8, mod, "Auth", "Non-SA blocked from all /superadmin/* pages", "P0", "Security", "Employer, Candidate",
           "Non-SA session", "1. Hit each SA route", "Denied", ""),
    ]
    m[mod] = cases

    # ---- 20 Uploads S3 ----
    mod = "20 Uploads S3"
    mid = "20"
    cases = [
        tc(mid, 1, mod, "Config", "Missing S3 config surfaces clear error (not silent success)", "P0", "Edge", "Candidate",
           "Temporarily unset AWS env locally", "1. Attempt photo upload", "Actionable error; no fake success", "Restore env after test; never blank prod secrets"),
        tc(mid, 2, mod, "Path", "Uploaded keys use internship-portal/ prefix", "P0", "Functional", "Employer",
           "Working S3", "1. Upload logo\n2. Inspect stored key/URL pattern", "Prefix internship-portal/…", ""),
        tc(mid, 3, mod, "Auth", "Upload endpoints require auth", "P0", "Security", "Public",
           "No session", "1. POST upload API directly", "401/403", ""),
    ]
    m[mod] = cases

    # ---- 21 Security Access ----
    mod = "21 Security Access"
    mid = "21"
    cases = [
        tc(mid, 1, mod, "RBAC", "Candidate APIs reject employer-only mutations", "P0", "Security", "Candidate",
           "Candidate session", "1. Call employer create internship API", "403", ""),
        tc(mid, 2, mod, "RBAC", "Employer APIs reject candidate-only apply-as-other", "P0", "Security", "Employer",
           "Employer session", "1. Attempt apply-as-candidate style call if any", "403/not allowed", ""),
        tc(mid, 3, mod, "IDOR", "Cannot modify another employer's internship by id", "P0", "Security", "Employer",
           "Foreign internship id", "1. PATCH/edit foreign id", "403/404", ""),
        tc(mid, 4, mod, "Session", "Mutating /api/ip/* without session fails", "P0", "Security", "Public",
           "No cookies", "1. POST offers/applications endpoints", "401/403", ""),
        tc(mid, 5, mod, "Tables", "App uses ip_* tables only (no writes to Placement Hub tables)", "P1", "Regression", "Developer / QA",
           "DB access", "1. Perform IP flows\n2. Confirm touched tables are ip_*", "No ism_*/PH table mutations from IP app", "README key decision"),
    ]
    m[mod] = cases

    # ---- 22 Mobile UI ----
    mod = "22 Mobile UI"
    mid = "22"
    cases = [
        tc(mid, 1, mod, "Layout", "Candidate home usable at 375px width", "P1", "UI", "Candidate",
           "Mobile viewport", "1. Open /candidate at 375px", "Primary nav/CTAs reachable; no clipped critical actions", ""),
        tc(mid, 2, mod, "Layout", "Employer posting form usable on mobile", "P1", "UI", "Employer",
           "Mobile viewport", "1. Open /employer/internships/new", "Fields and submit reachable", ""),
        tc(mid, 3, mod, "Layout", "Login + captcha usable on mobile", "P1", "UI", "Public",
           "375px", "1. Complete login", "Captcha + submit usable", ""),
    ]
    m[mod] = cases

    # ---- 23 Error Handling ----
    mod = "23 Error Handling"
    mid = "23"
    cases = [
        tc(mid, 1, mod, "404", "Unknown route shows not-found without app crash", "P1", "Functional", "Public",
           "None", "1. Open /this-route-does-not-exist-ip", "Not-found UI; app shell stable", ""),
        tc(mid, 2, mod, "API", "Malformed JSON body returns 400 not 500", "P1", "Negative", "Developer / QA",
           "Auth session optional", "1. POST bad JSON to an /api/ip endpoint", "400/422 class error", ""),
        tc(mid, 3, mod, "Empty", "Empty lists show empty state not spinner forever", "P2", "UI", "Candidate",
           "New candidate no apps", "1. Open applications", "Empty state copy", ""),
    ]
    m[mod] = cases

    # ---- 24 Public Content & Help ----
    mod = "24 Public Content & Help"
    mid = "24"
    cases = [
        tc(mid, 1, mod, "Landing", "Public landing / loads", "P0", "Functional", "Public",
           "None", "1. Open /", "Marketing/landing content; CTAs to login/register", ""),
        tc(mid, 2, mod, "Static", "How it works / guidelines / help pages load", "P1", "Functional", "Public",
           "None", "1. Open /how-it-works, /guidelines, /help", "Each returns 200 content", ""),
        tc(mid, 3, mod, "Nav", "Public pages link to login/register", "P2", "UI", "Public",
           "Landing", "1. Follow primary CTAs", "Reach auth/register", ""),
    ]
    m[mod] = cases

    # ---- 25 Sandbox Demo QA ----
    mod = "25 Sandbox Demo QA"
    mid = "25"
    cases = [
        tc(mid, 1, mod, "Demo logins", "Demo candidate/employer/SA passwords work", "P0", "Functional", "Public",
           "Seed/bootstrap applied", "1. Login each demo account", "Each reaches correct role home",
           f"{DEMO['candidate']}; {DEMO['employer']}; {DEMO['superadmin']}"),
        tc(mid, 2, mod, "Bootstrap", "/api/ip/bootstrap seeds SuperAdmin when used in allowed env", "P1", "Functional", "Developer / QA",
           "Dev/allowed env", "1. Run bootstrap per README", "superadmin@internship.local usable", "Do not expose bootstrap on locked prod without auth"),
        tc(mid, 3, mod, "Isolation", "IP demo data does not alter Placement Hub demo tenants", "P0", "Regression", "Developer / QA",
           "Shared Supabase", "1. Run IP flows\n2. Spot-check PH tables untouched", "Only ip_* changes", ""),
        tc(mid, 4, mod, "Env safety", "Deploy/docs never instruct blanking .env files", "P0", "Process", "Developer / QA",
           "Docs/scripts", "1. Review setup docs used in QA", "No truncate/wipe of env secrets", "Workspace safety rule"),
    ]
    m[mod] = cases

    return m


COVERAGE_ROWS = [
    # Capability, Candidate, Employer, SuperAdmin, Public
    ("Sign-in / session", "Y", "Y", "Y", "Y"),
    ("Register", "Y", "Y", "—", "Y"),
    ("Profile / photo-logo / docs", "Y", "Y", "—", "—"),
    ("Browse / apply internships", "Y (5 pts)", "—", "—", "—"),
    ("application_allowance column", "awarded/shown", "—", "—", "— NOT spent on apply"),
    ("Post / manage internships", "—", "Y (approved+complete)", "oversee", "—"),
    ("Applicants pipeline", "—", "Y", "—", "—"),
    ("Candidate search / invite", "opt-in", "Y", "—", "—"),
    ("Offers (respond)", "accept/decline", "create/list", "—", "—"),
    ("Messages", "Y", "Y", "Y", "—"),
    ("Points / referral /r", "Y", "Y", "—", "landing"),
    ("Points → post credits convert", "NO", "Y (50→1)", "—", "—"),
    ("Viral board / verify", "—", "Y", "Y", "—"),
    ("LinkedIn promotions", "—", "Y", "Y", "—"),
    ("Ratings / endorsements", "Y", "Y", "—", "—"),
    ("Feature ideas", "Y", "Y", "moderate", "read"),
    ("Notifications", "Y", "Y", "no SA page", "—"),
    ("Employer approvals / documents", "—", "submit", "Y", "—"),
    ("Login report / stats", "—", "—", "Y", "—"),
    ("S3 uploads", "photo", "logo/docs", "—", "—"),
    ("Captcha gate", "on /", "on /", "SA login", "Y"),
    ("/login route", "redirect→/", "redirect→/", "—", "Y"),
]


def style_header(ws, row=1):
    for col, name in enumerate(COLS, 1):
        cell = ws.cell(row=row, column=col, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def write_cases(ws, cases: list[dict], start_row: int = 2):
    for i, c in enumerate(cases):
        r = start_row + i
        values = [
            c["id"],
            c["module"],
            c["feature"],
            c["title"],
            c["priority"],
            c["type"],
            c["roles"],
            c["pre"],
            c["steps"],
            c["expected"],
            c["notes"],
            c["automation"],
            "Not Run",
            "",
            "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = [14, 28, 18, 48, 10, 12, 22, 36, 44, 44, 36, 12, 10, 24, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{start_row + len(cases) - 1}"


def priority_counts(cases: list[dict]):
    p0 = sum(1 for c in cases if c["priority"] == "P0")
    p1 = sum(1 for c in cases if c["priority"] == "P1")
    p2 = sum(1 for c in cases if c["priority"] in ("P2", "P3"))
    return len(cases), p0, p1, p2


def main():
    modules = build_cases()
    wb = Workbook()

    # Index
    idx = wb.active
    idx.title = "Index"
    idx["A1"] = "Internship Portal — Test Case Index"
    idx["A1"].font = INDEX_TITLE_FONT
    idx["A2"] = (
        "Authored for campus-placement-multiuser/internship-portal (ip_*). "
        "Same COLUMN schema as PlacementHub-Test-Cases.xlsx / campus-placement-multiuser/test_cases.csv, "
        "but cases are IP-only (points-on-apply, viral, referrals, ethics→profile_complete, SA ops). "
        "NOT a copy of PH modules (no drives/college/FCFS/alumni/student). "
        "Audited against running sibling internship-portal + mono source (2026-08). "
        "Known code quirks documented in cases (allowance display-only; /login→/; CSV export; draft API gate)."
    )
    idx["A2"].alignment = Alignment(wrap_text=True)
    idx.merge_cells("A2:G2")
    idx.row_dimensions[2].height = 48

    idx["A4"] = "#"
    idx["B4"] = "Module / Tab (click to open)"
    idx["C4"] = "Case Count"
    idx["D4"] = "P0"
    idx["E4"] = "P1"
    idx["F4"] = "P2+"
    idx["G4"] = "Link"
    for col in range(1, 8):
        idx.cell(row=4, column=col).fill = HEADER_FILL
        idx.cell(row=4, column=col).font = HEADER_FONT

    idx["A5"] = 0
    idx["B5"] = "Coverage Matrix"
    idx["C5"] = 0
    idx["D5"] = 0
    idx["E5"] = 0
    idx["F5"] = 0
    idx["G5"] = "Open >"
    idx["G5"].font = LINK_FONT
    idx["G5"].hyperlink = "#'Coverage Matrix'!A1"

    row = 6
    total = p0t = p1t = p2t = 0
    sheet_order = list(modules.keys())
    for n, mod in enumerate(sheet_order, 1):
        cases = modules[mod]
        count, p0, p1, p2 = priority_counts(cases)
        total += count
        p0t += p0
        p1t += p1
        p2t += p2
        idx.cell(row=row, column=1, value=n)
        idx.cell(row=row, column=2, value=mod)
        idx.cell(row=row, column=3, value=count)
        idx.cell(row=row, column=4, value=p0)
        idx.cell(row=row, column=5, value=p1)
        idx.cell(row=row, column=6, value=p2)
        link = idx.cell(row=row, column=7, value="Open >")
        link.font = LINK_FONT
        # Sheet title max 31 chars — our titles fit
        link.hyperlink = f"#'{mod}'!A1"
        row += 1

    idx.cell(row=row, column=2, value="TOTAL")
    idx.cell(row=row, column=3, value=total)
    idx.cell(row=row, column=4, value=p0t)
    idx.cell(row=row, column=5, value=p1t)
    idx.cell(row=row, column=6, value=p2t)
    row += 2
    idx.cell(row=row, column=1, value="Demo logins")
    idx.cell(row=row, column=2, value=f"Candidate: {DEMO['candidate']} | Employer: {DEMO['employer']} | SA: {DEMO['superadmin']}")
    row += 1
    idx.cell(row=row, column=1, value="Mail QA")
    idx.cell(row=row, column=2, value=DEMO["yopmail"])
    row += 1
    idx.cell(row=row, column=1, value="Priority legend")
    idx.cell(row=row, column=2, value="P0 = critical path/security/economy | P1 = high workflows | P2+ = polish/edge")
    row += 1
    idx.cell(row=row, column=1, value="Columns")
    idx.cell(row=row, column=2, value=", ".join(COLS))
    idx.column_dimensions["A"].width = 14
    idx.column_dimensions["B"].width = 36
    idx.column_dimensions["C"].width = 12
    idx.column_dimensions["G"].width = 10

    # Coverage Matrix
    cov = wb.create_sheet("Coverage Matrix")
    cov["A1"] = "<< Back to Index"
    cov["A1"].font = LINK_FONT
    cov["A1"].hyperlink = "#Index!A1"
    cov["A2"] = "Role × capability coverage (Internship Portal)"
    headers = ["Capability", "Candidate", "Employer", "SuperAdmin", "Public"]
    for col, h in enumerate(headers, 1):
        cell = cov.cell(row=3, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, r in enumerate(COVERAGE_ROWS):
        for col, val in enumerate(r, 1):
            cov.cell(row=4 + i, column=col, value=val)
    for i, w in enumerate([40, 14, 14, 14, 12], 1):
        cov.column_dimensions[get_column_letter(i)].width = w

    # Module sheets
    for mod, cases in modules.items():
        ws = wb.create_sheet(mod[:31])  # Excel limit
        ws["A1"] = "<< Back to Index"
        ws["A1"].font = LINK_FONT
        ws["A1"].hyperlink = "#Index!A1"
        # Put header on row 2 to mirror PH "back link then header" pattern visually,
        # but PH puts header as second nonempty after back — columns start row 2.
        # We'll put COLS on row 2 and cases from 3... Actually PH has back on row1, header row2, data row3.
        # Our style_header writes to row 1 by default — use row 2.
        for col, name in enumerate(COLS, 1):
            cell = ws.cell(row=2, column=col, value=name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        write_cases(ws, cases, start_row=3)
        # fix autofilter to include header row 2
        ws.auto_filter.ref = f"A2:{get_column_letter(len(COLS))}{2 + len(cases)}"
        ws.freeze_panes = "A3"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    OUT_COPY.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_COPY)
    if OUT_SIBLING.parent.parent.exists():
        OUT_SIBLING.parent.mkdir(parents=True, exist_ok=True)
        wb.save(OUT_SIBLING)
        print(f"Wrote {OUT_SIBLING}")
    print(f"Wrote {OUT}")
    print(f"Wrote {OUT_COPY}")
    print(f"Modules={len(modules)} cases={total} P0={p0t} P1={p1t} P2+={p2t}")


if __name__ == "__main__":
    main()
