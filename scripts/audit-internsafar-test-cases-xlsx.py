#!/usr/bin/env python3
"""Print full InternSafar-Test-Cases.xlsx status summary (all TC-IS rows).

Use after partial QA runs so reports always include workbook-wide totals,
not just the cases executed in the current session.

  python scripts/audit-internsafar-test-cases-xlsx.py
"""
from __future__ import annotations

import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "test-cases" / "InternSafar-Test-Cases.xlsx"
SKIP = frozenset({"Index", "Coverage", "Notes", "How to use"})


def audit(path: Path = XLSX) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    overall: Counter = Counter()
    by_auto: dict[str, Counter] = defaultdict(Counter)
    by_sheet: dict[str, Counter] = defaultdict(Counter)
    non_pass: list[dict] = []

    for ws in wb.worksheets:
        if ws.title in SKIP:
            continue
        hr = None
        cols: dict[str, int] = {}
        for r in range(1, 10):
            vals = [ws.cell(r, c).value for c in range(1, 20)]
            if vals and "TC ID" in vals and "Status" in vals:
                hr = r
                cols = {str(v): i + 1 for i, v in enumerate(vals) if v}
                break
        if not hr:
            continue
        for r in range(hr + 1, ws.max_row + 1):
            tc = ws.cell(r, cols["TC ID"]).value
            if not tc or not str(tc).startswith("TC-IS-"):
                continue
            status = (ws.cell(r, cols["Status"]).value or "Not Run").strip()
            auto_col = cols.get("Automation")
            automation = (
                (ws.cell(r, auto_col).value or "Unknown").strip() if auto_col else "Unknown"
            )
            overall[status] += 1
            by_auto[automation][status] += 1
            by_sheet[ws.title][status] += 1
            if status != "Pass":
                leg_col = cols.get("Legacy ID")
                legacy = ws.cell(r, leg_col).value if leg_col else ""
                non_pass.append(
                    {
                        "tcId": str(tc),
                        "sheet": ws.title,
                        "status": status,
                        "automation": automation,
                        "legacyId": str(legacy or "").strip(),
                    }
                )

    wb.close()
    total = sum(overall.values())
    return {
        "workbook": str(path),
        "totalTcIs": total,
        "byStatus": dict(overall),
        "byAutomation": {k: dict(v) for k, v in sorted(by_auto.items())},
        "bySheet": {k: dict(v) for k, v in sorted(by_sheet.items())},
        "nonPassCount": len(non_pass),
        "nonPass": non_pass,
    }


def main() -> None:
    report = audit()
    print(json.dumps(report, indent=2))
    s = report["byStatus"]
    total = report["totalTcIs"]
    passed = s.get("Pass", 0)
    print(
        f"\nSummary: {total} TC-IS rows | Pass {passed} | "
        f"Fail {s.get('Fail', 0)} | Blocked {s.get('Blocked', 0)} | "
        f"Not Run {s.get('Not Run', 0)} | other {total - passed - s.get('Fail', 0) - s.get('Blocked', 0) - s.get('Not Run', 0)}",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
