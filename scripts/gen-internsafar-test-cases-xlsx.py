# -*- coding: utf-8 -*-
"""InternSafar test-case workbook — Reference A columns/styling, InternSafar content."""
from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
ROOT = Path(__file__).resolve().parent.parent
DUMP = ROOT / "scripts" / "_ref_b_dump.json"
OUT = ROOT / "test-cases" / "InternSafar-Test-Cases.xlsx"

COLS = [
    "TC ID",
    "Module",
    "Feature",
    "Title",
    "Priority",
    "Type",
    "Role(s)",
    "Preconditions",
    "Test Steps",
    "Expected Result",
    "Test Data / Notes",
    "Automation",
    "Status",
    "Actual Result",
    "Executed At",
    "Legacy ID",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri")
INDEX_TITLE_FONT = Font(size=16, bold=True, color="1F4E79", name="Calibri")
LINK_FONT = Font(color="0563C1", underline="single", name="Calibri")
BODY_FONT = Font(name="Calibri", size=11)
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
PASS_FONT = Font(bold=True, color="006100", name="Calibri")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
FAIL_FONT = Font(bold=True, color="9C0006", name="Calibri")
BLOCK_FILL = PatternFill("solid", fgColor="D9D9D9")
BLOCK_FONT = Font(bold=True, color="595959", name="Calibri")
NOTRUN_FILL = PatternFill("solid", fgColor="FFF2CC")

DEMO = {
    "candidate": "lawsonlclintern+1@gmail.com / Admin@123",
    "candidate2": "lawsonlclintern+2@gmail.com / Admin@123",
    "candidate_hidden": "lawsonlclintern+3@gmail.com / Admin@123",
    "employer": "placementhubsupport@gmail.com / Admin@123",
    "employer_pending": "support+3@placementhub.online (pending)",
    "superadmin": "support@placementhub.online / Admin@123",
}

# Prefix → InternSafar module (current product, not Reference A 25-module list)
PREFIX_MODULE = {
    "PUB": ("01 Public & Landing", "01"),
    "HELP": ("01 Public & Landing", "01"),
    "AUTH": ("02 Auth & Access", "02"),
    "REG": ("03 Registration", "03"),
    "PERM": ("04 Permissions & Role", "04"),
    "ACCT": ("05 Account & Security", "05"),
    "CAND-P": ("06 Candidate Profile", "06"),
    "ACA": ("06 Candidate Profile", "06"),
    "CAND-X": ("06 Candidate Profile", "06"),
    "CAND-B": ("07 Browse Save Apply", "07"),
    "CAND-A": ("07 Browse Save Apply", "07"),
    "CAND-AP": ("08 Candidate Pipeline UX", "08"),
    "CAND-M": ("12 Messages", "12"),
    "CAND-O": ("11 Offers Ratings Completions", "11"),
    "CAND-R": ("13 Points Referrals Viral", "13"),
    "CAND-N": ("12 Messages", "12"),
    "CAND-D": ("08 Candidate Pipeline UX", "08"),
    "OFF-R": ("11 Offers Ratings Completions", "11"),
    "EMP-P": ("09 Employer Postings Pipeline", "09"),
    "EMP-H": ("09 Employer Postings Pipeline", "09"),
    "EMP-I": ("09 Employer Postings Pipeline", "09"),
    "EMP-PL": ("09 Employer Postings Pipeline", "09"),
    "EMP-C": ("10 Employer Candidate Search", "10"),
    "EMP-M": ("12 Messages", "12"),
    "EMP-O": ("11 Offers Ratings Completions", "11"),
    "EMP-AN": ("09 Employer Postings Pipeline", "09"),
    "EMP-V": ("13 Points Referrals Viral", "13"),
    "EMP-R": ("13 Points Referrals Viral", "13"),
    "EMP-N": ("12 Messages", "12"),
    "COMP": ("11 Offers Ratings Completions", "11"),
    "SA": ("14 SuperAdmin Ops", "14"),
    "IDEA": ("15 Feature Ideas", "15"),
    "PTS": ("13 Points Referrals Viral", "13"),
    "FILE": ("17 Files & Uploads", "17"),
    "RATE": ("11 Offers Ratings Completions", "11"),
    "ERR": ("18 Cross-cutting", "18"),
    "MAIL": ("18 Cross-cutting", "18"),
    "SHELL": ("18 Cross-cutting", "18"),
    "REGX": ("18 Cross-cutting", "18"),
    "NAV": ("18 Cross-cutting", "18"),
    "BOOT": ("18 Cross-cutting", "18"),
}

CORRECTED = {
    "PUB-1": "Home sign-in uses #email and #password.",
    "PUB-2": "/login redirects to `/` and keeps query params.",
    "AUTH-1": "Candidate/employer: Gmail+ cores on `/` with captcha.",
    "AUTH-15": "SuperAdmin: support@placementhub.online on /superadmin/login (#sa-email, #sa-password).",
    "BOOT-1": "Bootstrap keeps support@placementhub.online and the Gmail+ showcase users.",
    "CAND-P-1": "Profile is five tabs: Basics, Academic, Work Readiness, Privacy & Photo, Endorsements (read-only).",
    "CAND-P-2": "Photo/privacy on tab 4; academics via /api/ip/candidate/academics.",
    "CAND-P-3": "Incomplete-profile banner on /candidate and /candidate/profile.",
    "ACA-1": "Academic rows save through the academics API.",
    "CAND-X-1": "Export: GET /api/ip/candidate/export from profile.",
    "CAND-B-2": "Browse last-used filters: tableKey candidate.internships.",
    "CAND-AP-1": "Applications list: tableKey candidate.applications.",
    "CAND-N-1": "Notifications page: /candidate/notifications.",
    "EMP-C-1": "Search then full page /employer/candidates/[id].",
    "EMP-C-2": "Search filters: tableKey employer.candidates.",
    "EMP-C-3": "Invite requires internshipId.",
    "EMP-C-4": "Cannot invite on another employer’s posting.",
    "EMP-I-2": "Publish costs 50 points; draft does not.",
    "PTS-2": "No convert-points UI or API. Apply costs 5; publish costs 50.",
    "SHELL-1": "Only candidate and employer have a Notifications nav item.",
    "REGX-1": "Register uses real Google OAuth verification; login on `/` is email/password only.",
    "AUTH-4": "Captcha must be solved on home login.",
    "FILE-1": "Files go through /api/ip/files.",
    "AUTH-17": "Account password change requires upper + digit + special.",
    "AUTH-18": "Weak new password on /account is rejected.",
    "REG-C-1": "Gmail path: Sign up with Google → gv token → temp password emailed; sign in on `/`.",
    "REG-C-4": "Form path (API): password length ≥ 8 only; account stays pending until SuperAdmin.",
    "EMP-N-1": "Employer notifications: /employer/notifications.",
}

# How InternSafar works today.
LIVE = {
    "PUB-1": {
        "title": "Sign-in is on the home page",
        "steps": (
            "1. Open http://localhost:3000/ signed out.\n"
            "2. Confirm Email (id=email), Password (id=password), captcha, Remember this device, and Submit.\n"
            "3. Confirm links to Register, How it works, Guidelines, Help."
        ),
        "expected": "The login form is on `/`. Fields are #email and #password. There is no separate product login screen.",
    },
    "PUB-2": {
        "title": "/login sends you to home and keeps the query string",
        "steps": (
            "1. Signed out, open /login.\n"
            "2. Open /login?email=lawsonlclintern%2B1%40gmail.com&next=/candidate.\n"
            "3. Watch the address bar."
        ),
        "expected": "Browser ends on `/?…` (or `/`). You never stay on a /login form. Email query can prefill #email on home.",
    },
    "AUTH-1": {
        "title": "Valid email + password + captcha lands on that role’s home",
        "steps": (
            "1. Open `/`.\n"
            "2. Candidate: lawsonlclintern+1@gmail.com / Admin@123 + captcha → submit.\n"
            "3. Sign out. Repeat as employer placementhubsupport@gmail.com / Admin@123.\n"
            "4. SuperAdmin uses /superadmin/login with #sa-email / #sa-password, not the home form."
        ),
        "expected": "Candidate → /candidate. Employer → /employer. SuperAdmin from /superadmin/login → /superadmin. Home form is for candidate/employer.",
    },
    "AUTH-4": {
        "title": "Wrong or empty captcha blocks login on home",
        "steps": (
            "1. On `/` enter a valid candidate email and password.\n"
            "2. Leave captcha blank and submit.\n"
            "3. Enter a wrong captcha and submit.\n"
            "4. Solve captcha correctly and submit."
        ),
        "expected": "Blank/wrong captcha: error, stay signed out. Correct captcha: session starts and /candidate (or role home).",
    },
    "AUTH-15": {
        "title": "SuperAdmin signs in only on /superadmin/login",
        "steps": (
            "1. Open /superadmin/login.\n"
            "2. Sign in with support@placementhub.online / Admin@123 + captcha (#sa-email, #sa-password).\n"
            "3. On `/` try the same SuperAdmin email in #email/#password.\n"
            "4. On /superadmin/login try lawsonlclintern+1@gmail.com."
        ),
        "expected": "SA page accepts SuperAdmin and goes to /superadmin. Candidate email on SA page is rejected. SuperAdmin is not meant to use the public home form as their login.",
    },
    "AUTH-17": {
        "title": "Signed-in user can change password from Account",
        "steps": (
            "1. Sign in as candidate on `/` (#email/#password).\n"
            "2. Open /account.\n"
            "3. Current password Admin@123. New password must be ≥8 and include uppercase, a digit, and a special character (e.g. Admin@124).\n"
            "4. Sign in again with the new password, then restore Admin@123 if this is a shared core account."
        ),
        "expected": "Change succeeds only when the new password meets Account rules. Register-form rule (length only) does not apply here.",
    },
    "AUTH-18": {
        "title": "Wrong current password or weak new password is rejected",
        "steps": (
            "1. On /account, submit wrong current password.\n"
            "2. Submit new password of 7 characters.\n"
            "3. Submit 8+ lowercase-only (no upper/digit/special)."
        ),
        "expected": "All three fail. Account change-password requires 8+ and uppercase + number + special. Form signup only checks length ≥ 8 — that is a different screen.",
    },
    "BOOT-1": {
        "title": "App bootstrap keeps SuperAdmin seed",
        "steps": (
            "1. POST /api/ip/bootstrap (happens on SA login page load too).\n"
            "2. Sign in support@placementhub.online / Admin@123 on /superadmin/login.\n"
            "3. Confirm core candidate/employer are the Gmail+ accounts, not deleted by bootstrap."
        ),
        "expected": "SuperAdmin can sign in. Showcase accounts stay lawsonlclintern+1@gmail.com and placementhubsupport@gmail.com.",
    },
    "PTS-2": {
        "title": "There is no convert-points-to-credits action",
        "steps": (
            "1. As candidate and employer, open Refer & earn / dashboard.\n"
            "2. Confirm no Convert / posting-credits button.\n"
            "3. POST /api/ip/points/convert if you want to probe the URL."
        ),
        "expected": "UI has no convert. That route is not in the app (404). Points are spent directly on apply (5) and publish (50).",
    },
    "REGX-2": {
        "title": "InternSafar only writes its own ip_* data",
        "steps": (
            "1. Run a normal InternSafar flow (login, apply or post).\n"
            "2. Confirm app APIs under /api/ip/ are used."
        ),
        "expected": "InternSafar pages and /api/ip/* succeed.",
    },
    "REGX-1": {
        "title": "Google OAuth on register verifies identity; login is email/password only",
        "steps": (
            "1. Open /register/candidate and /register/employer.\n"
            "2. Confirm real Sign up / Continue with Google (opens Google account chooser).\n"
            "3. Open `/` login — confirm there is no Google button.\n"
            "4. Attempt Google sign-in from `/` without registration intent → /?error=GoogleLoginDisabled."
        ),
        "expected": (
            "Register uses real Google OAuth (gv token handoff). "
            "Home sign-in is #email/#password only; Google never creates a portal login session."
        ),
    },
    "REG-C-1": {
        "title": "Gmail candidate signup emails a temp password and can sign in on home",
        "steps": (
            "1. Open /register/candidate signed out.\n"
            "2. Click Sign up with Google; complete consent with a unique unused @gmail.com.\n"
            "3. Confirm Registration complete and temp password email.\n"
            "4. Sign in on `/` with #email/#password using the emailed temp password."
        ),
        "expected": (
            "Account active immediately (registration_source=google). "
            "Login on `/` is email/password — not Google OAuth session."
        ),
    },
    "REG-C-4": {
        "title": "Form-path signup (API): university, year, password ≥8, captcha; stays pending",
        "steps": (
            "1. POST /api/ip/auth/register-candidate path=form (no UI on /register/candidate).\n"
            "2. Password 7 characters → rejected.\n"
            "3. Password 8+ characters + college + year + captcha → pending account.\n"
            "4. Try signing in on `/` before SuperAdmin approval."
        ),
        "expected": (
            "7 chars fail. 8+ creates pending account. Login blocked until SuperAdmin approves. "
            "Account change-password later is stricter (upper+digit+special)."
        ),
    },
    "SHELL-1": {
        "title": "Sidebar matches the signed-in role",
        "steps": (
            "1. Candidate: confirm nav includes Notifications → /candidate/notifications.\n"
            "2. Employer: Notifications → /employer/notifications.\n"
            "3. SuperAdmin: confirm nav from Dashboard through Feature ideas + Account — no Notifications item.\n"
            "4. Mobile drawer / collapse if present."
        ),
        "expected": "Labels match src/lib/ipNav.js. SuperAdmin has no notifications page. Candidate/employer do.",
    },
    "CAND-N-1": {
        "title": "Candidate notifications page filters, search, mark read",
        "steps": (
            "1. Sign in candidate on `/`.\n"
            "2. Open /candidate/notifications.\n"
            "3. Use filter chips, search, mark read, presets bar if shown."
        ),
        "expected": "Page loads. Filters work. SuperAdmin does not have this page.",
    },
    "EMP-N-1": {
        "title": "Employer notifications mailbox",
        "steps": "1. Sign in employer on `/`.\n2. Open /employer/notifications.\n3. Mark read / filter if present.",
        "expected": "Page loads for employer. SuperAdmin has no equivalent nav page.",
    },
}

# Risk: likelihood(1-3) x impact(1-4). P0 if score>=6 else P1 if >=4 else P2 unless B said High.
HIGH_RISK_PREFIX = ("AUTH", "PERM", "REG-C", "REG-E", "CAND-A", "EMP-I", "EMP-PL", "SA-A", "SA-F")


def module_for(legacy_id: str) -> tuple[str, str]:
    if legacy_id.startswith("CAND-AP"):
        return PREFIX_MODULE["CAND-AP"]
    if legacy_id.startswith("CAND-P"):
        return PREFIX_MODULE["CAND-P"]
    if legacy_id.startswith("CAND-X"):
        return PREFIX_MODULE["CAND-X"]
    if legacy_id.startswith("CAND-B"):
        return PREFIX_MODULE["CAND-B"]
    if legacy_id.startswith("CAND-A"):
        return PREFIX_MODULE["CAND-A"]
    if legacy_id.startswith("CAND-M"):
        return PREFIX_MODULE["CAND-M"]
    if legacy_id.startswith("CAND-O"):
        return PREFIX_MODULE["CAND-O"]
    if legacy_id.startswith("CAND-R"):
        return PREFIX_MODULE["CAND-R"]
    if legacy_id.startswith("CAND-N"):
        return PREFIX_MODULE["CAND-N"]
    if legacy_id.startswith("CAND-D"):
        return PREFIX_MODULE["CAND-D"]
    if legacy_id.startswith("EMP-PL"):
        return PREFIX_MODULE["EMP-PL"]
    if legacy_id.startswith("EMP-AN"):
        return PREFIX_MODULE["EMP-AN"]
    if legacy_id.startswith("OFF-R"):
        return PREFIX_MODULE["OFF-R"]
    if legacy_id.startswith("REGX"):
        return PREFIX_MODULE["REGX"]
    if legacy_id.startswith("REG"):
        return PREFIX_MODULE["REG"]
    prefix = legacy_id.split("-")[0]
    if prefix in PREFIX_MODULE:
        return PREFIX_MODULE[prefix]
    return ("18 Cross-cutting", "18")


def pri_from_b(row: dict, legacy_id: str) -> str:
    sev = str(row.get("severity") or "").lower()
    if any(legacy_id.startswith(p) for p in HIGH_RISK_PREFIX):
        if "low" in sev:
            return "P1"
        return "P0"
    if "high" in sev:
        return "P0"
    if "low" in sev:
        return "P2"
    return "P1"


def type_from_b(t: str) -> str:
    t = (t or "Functional").strip()
    mapping = {
        "Functional": "Functional",
        "Negative": "Negative",
        "Edge Case": "Edge",
        "Error Handling": "Negative",
        "UI Validation": "UI",
        "Regression": "Regression",
        "To Verify": "Functional",
        "Security": "Security",
        "Accessibility": "UI",
    }
    return mapping.get(t, t if t else "Functional")


def roles_guess(legacy_id: str, module: str) -> str:
    if legacy_id.startswith(("PUB", "HELP", "REG-C", "REG-E", "REGX")):
        return "Public"
    if legacy_id.startswith("AUTH"):
        if "15" in legacy_id or "22" in legacy_id:
            return "SuperAdmin, Public"
        return "Public, Candidate, Employer"
    if legacy_id.startswith("SA") or legacy_id.startswith("BOOT"):
        return "SuperAdmin"
    if legacy_id.startswith("PERM"):
        return "Public, Candidate, Employer, SuperAdmin"
    if legacy_id.startswith("EMP") or legacy_id.startswith("COMP"):
        return "Employer"
    if legacy_id.startswith("CAND") or legacy_id.startswith("ACA"):
        return "Candidate"
    if legacy_id.startswith("IDEA"):
        return "Candidate, Employer, SuperAdmin"
    if legacy_id.startswith(("FILE", "PTS", "RATE", "MAIL", "SHELL", "NAV", "ERR")):
        return "Candidate, Employer, SuperAdmin"
    return "As applicable"


def methodology_notes(legacy_id: str, row: dict, recon: str) -> str:
    bits = [
        f"Legacy ID: {legacy_id} (plain text map only; Reference B is not a runtime dependency).",
        f"Reconciliation: {recon}.",
        "QASkills test-plan: EP + BVA + decision table + risk (L×I). Status reset: Not Run.",
        f"Seed: Candidate {DEMO['candidate']}; Employer {DEMO['employer']}; SA {DEMO['superadmin']}. No real PII; Gmail+ aliases are product demo identities.",
        "InternSafar logins: candidate/employer on `/` (#email, #password); SuperAdmin on /superadmin/login (#sa-email, #sa-password).",
    ]
    if legacy_id in CORRECTED:
        bits.append("Correction vs old checklist: " + CORRECTED[legacy_id])
    summary = str(row.get("summary") or "")
    if "password" in summary.lower() or legacy_id.startswith(("AUTH-17", "AUTH-18", "REG-C-4")):
        bits.append(
            "BVA password: form register min 8; change-password min 8 + 1 upper + 1 digit + 1 special. Cases: 7, 8, 9 chars; missing class."
        )
    if "email" in summary.lower() or legacy_id.startswith("REG-C"):
        bits.append(
            "EP emails: user@gmail.com (valid), user@googlemail.com (accepted), user@yahoo.com (reject), user@gmail.com.evil (reject), empty, no @."
        )
    if "captcha" in summary.lower() or legacy_id in ("AUTH-4", "AUTH-10", "REGX-3"):
        bits.append("EP captcha: correct, blank, wrong, stale token, missing token.")
    if legacy_id.startswith("CAND-A") or legacy_id in ("PTS-1",):
        bits.append("Economy DT: apply costs 5 points; allowance display must not be confused with spend. EP balance 0, 4, 5, 6.")
    if legacy_id.startswith("EMP-I") and "50" in summary:
        bits.append("Publish DT: points>=50 × approved × profile_complete × not already published.")
    bits.append(f"Product reference: {row.get('reference') or 'InternSafar sibling src/'}")
    return "\n".join(bits)


def numbered_steps(desc: str) -> str:
    text = str(desc or "").strip()
    if not text:
        return "1. Follow the feature on the current InternSafar route.\n2. Capture request/response if API-backed.\n3. Reload and confirm persistence."
    if text[0].isdigit() and ". " in text[:4]:
        return text
    parts = [p.strip() for p in text.replace("\r", "").split("\n") if p.strip()]
    if len(parts) == 1:
        return "1. " + parts[0] + "\n2. Observe UI and network.\n3. Confirm expected persistence / error."
    return "\n".join(f"{i}. {p.lstrip('0123456789. ')}" for i, p in enumerate(parts, 1))


def tc_row(**kwargs):
    return kwargs


def rewrite_from_b(row: dict, seq: dict) -> dict:
    lid = row["id"]
    mod, mid = module_for(lid)
    recon = "Corrected/rewritten" if lid in CORRECTED else "Carried forward (rewritten into new schema)"
    n = seq[mid]
    seq[mid] += 1
    live = LIVE.get(lid) or {}
    title = live.get("title") or str(row.get("summary") or lid).strip()
    feature = str(row.get("section") or "General")
    steps = live.get("steps") or numbered_steps(row.get("description"))
    expected = live.get("expected") or str(
        row.get("expected") or "Behaviour matches InternSafar as implemented today."
    )
    return tc_row(
        id=f"TC-IS-{mid}-{n:03d}",
        module=mod,
        feature=feature[:40],
        title=title,
        priority=pri_from_b(row, lid),
        type=type_from_b(row.get("type")),
        roles=roles_guess(lid, mod),
        pre=(
            "InternSafar running (internship-portal, npm run dev). "
            f"Candidate {DEMO['candidate']}. Employer {DEMO['employer']}. SuperAdmin {DEMO['superadmin']}. "
            "Home login: #email / #password + captcha."
        ),
        steps=steps,
        expected=expected,
        notes=methodology_notes(lid, row, recon),
        automation="Manual (sibling has scripts/run-ip-checklist-qa.mjs coverage for many AUTH/PERM IDs — must be remapped to TC-IS-* before treating as green)",
        legacy=lid,
        recon=recon,
    )


def extra_cases(seq: dict) -> list[dict]:
    """Net-new InternSafar surfaces Reference B never listed."""
    out = []

    def add(mid, module, feature, title, pri, typ, roles, pre, steps, exp, notes, auto="Manual"):
        n = seq[mid]
        seq[mid] += 1
        out.append(
            tc_row(
                id=f"TC-IS-{mid}-{n:03d}",
                module=module,
                feature=feature,
                title=title,
                priority=pri,
                type=typ,
                roles=roles,
                pre=pre,
                steps=steps,
                expected=exp,
                notes=notes,
                automation=auto,
                legacy="",
                recon="New (not in Reference B)",
            )
        )

    seed = f"Candidate {DEMO['candidate']}; Employer {DEMO['employer']}"
    lists = [
        ("candidate.internships", "/candidate/internships", "Candidate"),
        ("candidate.applications", "/candidate/applications", "Candidate"),
        ("candidate.offers", "/candidate/offers", "Candidate"),
        ("candidate.messages", "/candidate/messages", "Candidate"),
        ("candidate.notifications", "/candidate/notifications", "Candidate"),
        ("candidate.referral", "/candidate/referral", "Candidate"),
        ("employer.internships", "/employer/internships", "Employer"),
        ("employer.candidates", "/employer/candidates", "Employer"),
        ("employer.offers", "/employer/offers", "Employer"),
        ("employer.messages", "/employer/messages", "Employer"),
        ("employer.notifications", "/employer/notifications", "Employer"),
        ("employer.referral", "/employer/referral", "Employer"),
    ]
    add(
        "16",
        "16 List Filters Sort Presets",
        "Last-used prefs",
        "Changing filters/sort persists after reload (debounced PUT /api/ip/table-filter-prefs)",
        "P0",
        "Functional",
        "Candidate, Employer",
        seed,
        "1. Open each list in the matrix.\n2. Change a filter and sort.\n3. Wait >450ms.\n4. Reload.",
        "Same filter/sort restored from ip_table_filter_prefs for that tableKey. SuperAdmin lists are out of scope (API requires candidate|employer).",
        "EP: empty filters vs populated. BVA debounce 450ms: change and reload immediately vs after wait. Risk 3×3=9 (wrong list = missed applicants).\n"
        + "tableKeys: "
        + ", ".join(t[0] for t in lists),
    )
    add(
        "16",
        "16 List Filters Sort Presets",
        "Default preset wins",
        "Named default preset overrides last-used prefs on load",
        "P0",
        "Functional",
        "Employer",
        DEMO["employer"],
        "1. On /employer/candidates set filters A and save preset 'View A' as default.\n2. Change filters to B (last-used).\n3. Reload.",
        "Default preset A is applied, not last-used B (useListPrefsSync: def wins over prefRes).",
        "Decision table: default exists Y/N × last-used exists Y/N. Isolation: employer.candidates vs employer.internships presets must not mix.",
    )
    add(
        "16",
        "16 List Filters Sort Presets",
        "Max 5 presets",
        "Sixth named preset is rejected with clear error",
        "P1",
        "Negative",
        "Candidate, Employer",
        seed,
        "1. Save 5 uniquely named presets on one tableKey.\n2. Attempt a 6th.",
        "400: already have 5 saved views. UI shows presetError. Existing five unchanged.",
        "BVA: 0, 1, 5, 6 presets. Duplicate name on same tableKey → unique constraint message.",
    )
    add(
        "16",
        "16 List Filters Sort Presets",
        "Duplicate name",
        "Same preset name on same tableKey is rejected; same name on another tableKey is allowed",
        "P1",
        "Negative",
        "Employer",
        DEMO["employer"],
        "1. Save 'Pipeline' on employer.internships.\n2. Save 'Pipeline' again on same key.\n3. Save 'Pipeline' on employer.candidates.",
        "Step 2 errors; step 3 succeeds (unique is user+table+name).",
        "EP partitions: duplicate vs unique vs whitespace-only name (Name required).",
    )
    add(
        "16",
        "16 List Filters Sort Presets",
        "Authz",
        "Anonymous and SuperAdmin cannot GET/POST list-presets or table-filter-prefs",
        "P0",
        "Security",
        "Public, SuperAdmin",
        "No session; then SA session",
        "1. GET /api/ip/list-presets?tableKey=employer.candidates without cookie.\n2. Repeat as SuperAdmin.",
        "401 unauthenticated; SuperAdmin is not in requireSession(['employer','candidate']) so 403.",
        "Missing tableKey → error. IDOR: cannot PATCH another user's preset id.",
    )
    add(
        "16",
        "16 List Filters Sort Presets",
        "Apply/delete",
        "Apply preset restores filters; delete removes it; toggling default is exclusive",
        "P1",
        "Functional",
        "Candidate",
        DEMO["candidate"],
        "1. Save two presets on candidate.internships.\n2. Set first as default.\n3. Set second as default.\n4. Apply first.\n5. Delete first.",
        "Only one is_default true. Apply changes visible rows. Delete 200 and list count drops.",
        "PATCH without id → error. DELETE missing id → error.",
    )
    add(
        "16",
        "16 List Filters Sort Presets",
        "Applicants pipeline key",
        "Applicant grid on /employer/internships/[id] uses per-posting tableKey and saved-views API",
        "P0",
        "Functional",
        "Employer",
        DEMO["employer"] + " with a Core Showcase posting that has applicants",
        "1. Open posting applicants.\n2. Filter/sort/save preset.\n3. Open a different posting.",
        "Prefs are scoped per internship id (employer.applicants.{id} / saved-views pipelineKey). Cross-posting leak is a fail.",
        "Also exercise GET/POST /api/ip/employer/saved-views. Risk 3×3=9.",
    )

    add(
        "06",
        "06 Candidate Profile",
        "Tab save isolation",
        "Each profile tab save persists only that section; switching tabs does not discard unsaved sibling edits without warning or save",
        "P0",
        "Functional",
        "Candidate",
        DEMO["candidate"],
        "1. Open /candidate/profile.\n2. Edit Basics, save.\n3. Edit Academic rows, save.\n4. Edit Work Readiness enums (Remote/Hybrid/On-site, commitment options).\n5. Privacy: searchable off, hide phone, photo show toggle.\n6. Endorsements tab is read-only.",
        "Values persist on reload per /api/ip/candidate/profile and /academics. Endorsements cannot be forged from UI. Employer search hides candidate when searchable off (unless they applied).",
        "EP work_mode: Remote/Hybrid/On-site. Commitment: none/other_internship/offline_classes/part_time_work/other/empty. Countries: India plus listed options. BVA CGPA 0, 10, 10.1 if validated.",
    )
    add(
        "06",
        "06 Candidate Profile",
        "Email change from profile",
        "New login email request + 6-digit verify from profile tab",
        "P1",
        "Functional",
        "Candidate",
        DEMO["candidate"] + "; QA mail override if used",
        "1. Enter a unique unused @gmail.com in Change Login Email.\n2. Request code.\n3. Submit wrong code.\n4. Submit correct code (or abort if you must not mutate core email — use a gen-accounts user).",
        "Wrong code rejected. Core account email must not be permanently changed in shared QA without a restore plan.",
        "Prefer generate:ip-test-data --mode=gen-accounts user. EP: non-gmail, duplicate email.",
    )
    add(
        "10",
        "10 Employer Candidate Search",
        "Full-page profile",
        "Employer candidate detail page loads discovery fields, optional application extras, notes, timeline, reminder",
        "P0",
        "Functional",
        "Employer",
        DEMO["employer"] + "; searchable candidate from search or applicant",
        "1. Open /employer/candidates.\n2. Open a candidate → /employer/candidates/{id}.\n3. Add a note.\n4. Set a reminder.\n5. Confirm phone hidden until shortlist rule allows.",
        "GET /api/ip/employer/candidates/{id} 200 when searchable or applicant. Phone masked when hide_phone_until_shortlist and status not allowed. Notes POST /applications/{id}/notes. Reminder POST /api/ip/employer/reminders.",
        "Security: guessing another org's candidate id without searchable/application → 404. applicationId query must be owned.",
    )
    add(
        "09",
        "09 Employer Postings Pipeline",
        "Bulk applicants",
        "Bulk applicant actions respect owned internship and allowed statuses",
        "P0",
        "Functional",
        "Employer",
        DEMO["employer"],
        "1. Open applicants for own posting.\n2. Bulk update via /api/ip/employer/internships/{id}/applicants/bulk.\n3. Repeat with a foreign internship id.",
        "Owned bulk succeeds; foreign id fails. Partial invalid ids do not silently update others' data.",
        "Decision table: empty list, mixed valid/invalid ids, closed posting.",
    )
    add(
        "09",
        "09 Employer Postings Pipeline",
        "Rejection templates",
        "CRUD rejection templates on /employer/rejection-templates",
        "P1",
        "Functional",
        "Employer",
        DEMO["employer"],
        "1. Create template.\n2. Use on reject if UI wires it.\n3. Delete.",
        "API /api/ip/employer/rejection-templates persists. Candidate cannot access.",
        "EP empty name vs long name. Authz employer-only.",
    )
    add(
        "09",
        "09 Employer Postings Pipeline",
        "Export jobs",
        "Applicant export job creates CSV/zip without hidden phones",
        "P1",
        "Functional",
        "Employer",
        DEMO["employer"] + " posting with applicants",
        "1. Trigger export from applicants UI.\n2. Poll /api/ip/employer/export-jobs/{id}.\n3. Open artifact.",
        "Export matches ipApplicantExport rules: hidden phones omitted. Cron /api/ip/cron/export-jobs is not a user toy — do not expose without secret.",
        "NFR: large applicant set does not hang the tab (job async).",
    )
    add(
        "09",
        "09 Employer Postings Pipeline",
        "Closure summary",
        "Closed posting shows closure summary API data",
        "P2",
        "Functional",
        "Employer",
        "Employer with a closable posting",
        "1. Close posting.\n2. GET /api/ip/employer/internships/{id}/closure-summary.",
        "Summary returns; candidates no longer see listing as applyable (CAND-B-1).",
        "Trace to CAND-B-1.",
    )
    add(
        "09",
        "09 Employer Postings Pipeline",
        "Employer lists",
        "Saved candidate lists API /api/ip/employer/lists",
        "P2",
        "Functional",
        "Employer",
        DEMO["employer"],
        "1. Create/list/delete a list if UI exposes it; else API-only.",
        "Employer-only. Candidate 403.",
        "If UI is missing, mark Actual as API-only when executing — do not invent a page.",
    )
    add(
        "17",
        "17 Files & Uploads",
        "Message attachment",
        "Thread attachment upload/download for participants only",
        "P1",
        "Security",
        "Candidate, Employer",
        seed,
        "1. Attach a small PDF/image on /candidate/messages/{id} or employer equivalent.\n2. Download as the other participant.\n3. Hit attachment URL as a third user.",
        "Participants can fetch /api/ip/messages/threads/{id}/attachment. Third party 403/404.",
        "EP file types; BVA size vs server limit.",
    )
    add(
        "14",
        "14 SuperAdmin Ops",
        "No SA notifications page",
        "SuperAdmin menu has no Notifications page",
        "P2",
        "Functional",
        "SuperAdmin",
        DEMO["superadmin"],
        "1. Sign in on /superadmin/login.\n2. Walk the left nav.\n3. Confirm there is no Notifications item (Messages exists).",
        "Product today: SuperAdmin uses Messages oversight only. Candidate/employer use /candidate/notifications and /employer/notifications. Do not look for a SuperAdmin notifications screen.",
        "GET /api/ip/notifications still accepts a SuperAdmin session (returns that user’s rows). There is no SA UI for it.",
    )
    add(
        "18",
        "18 Cross-cutting",
        "Ref catalogs",
        "City and degree reference endpoints serve pickers",
        "P2",
        "Functional",
        "Candidate, Employer",
        seed,
        "1. Open profile/posting forms that use city/degree.\n2. GET /api/ip/ref/cities and /api/ip/ref/degrees.",
        "Lists load; invalid search returns empty not 500.",
        "Auth: confirm whether public or session-gated in current route.js.",
    )
    add(
        "18",
        "18 Cross-cutting",
        "Nav badges",
        "Unread badges update after reading messages/notifications",
        "P1",
        "Functional",
        "Candidate, Employer",
        seed,
        "1. Note /api/ip/nav-badges counts.\n2. Read a message/notification.\n3. Navigate.",
        "Counts drop without full logout. SA badges if any match API.",
        "Related NAV-1.",
    )
    add(
        "02",
        "02 Auth & Access",
        "Login decision table",
        "Login DT: exists × password × captcha × active × 2FA × role landing",
        "P0",
        "Functional",
        "Public",
        "Known active candidate; pending form user; SA",
        "1. On `/` use #email/#password + captcha: valid candidate; wrong password; unknown email; pending form candidate; 2FA-enabled user.\n2. SuperAdmin: only /superadmin/login with #sa-email/#sa-password.",
        "Candidate → /candidate, employer → /employer from home. SuperAdmin → /superadmin from SA login. Pending form user cannot use home login until approved.",
        "InternSafar home fields are #email and #password.",
    )
    add(
        "03",
        "03 Registration",
        "Password BVA form path",
        "Form-path candidate password length 7 rejected, 8 accepted if other required fields present",
        "P0",
        "Negative",
        "Public",
        "/register/candidate form path",
        "1. Form path: Gmail + college + year + captcha + 7-character password.\n2. Same with 8-character password (no extra character classes required on signup).\n3. After approval, on /account try changing password to 8 lowercase letters only.",
        "Signup: 7 fails, 8+ length is enough on form path (pending until SA). Account change-password: 8 lowercase fails — needs uppercase, number, and special. Test both screens as they are.",
        "Two rules in InternSafar today: register form = length ≥ 8; /account change = length ≥ 8 + upper + digit + special.",
    )
    add(
        "07",
        "07 Browse Save Apply",
        "Skill match BVA",
        "Min match filter 0/1/100 and internships with empty eligibility skills",
        "P1",
        "Edge",
        "Candidate",
        DEMO["candidate"],
        "1. Set min match 0, 50, 100.\n2. Confirm empty eligibility internships behave as 100% advisory (CAND-A-6).",
        "Filter matches computeValidationScore / skillMatchPercent. No 500 on missing skills JSON.",
        "EP match percent.",
    )
    add(
        "11",
        "11 Offers Ratings Completions",
        "Rating bounds",
        "Stars 1 and 5 accepted; 0 and 6 rejected after completion only",
        "P1",
        "Negative",
        "Candidate, Employer",
        "Completed application (COMP path)",
        "1. Rate 0, 1, 5, 6 via POST /api/ip/ratings with internshipId on a hired/completed application.\n2. Rate with stars 5 and no internshipId (RATE-2).\n3. Rate while application is only applied/shortlisted.",
        "0 and 6 rejected. Missing internshipId rejected. Applied/shortlisted rejected until hired or completed. 1 and 5 accepted only after that gate (or 409 if that pair already rated that internship).",
        "BVA 0,1,5,6. Decision table: internshipId Y/N × application status applied|hired|completed. Unique (from,to,internship) → 409.",
    )
    add(
        "15",
        "15 Feature Ideas",
        "Sort partitions",
        "Ideas sort most_voted / newest / recently_updated and category filter",
        "P2",
        "Functional",
        "Candidate, Employer",
        seed,
        "1. Open /ideas.\n2. Switch sorts and category chips.\n3. Submit missing category (IDEA-5).",
        "Client sort matches page.js. SuperAdmin moderation on /superadmin/feature-ideas.",
        "EP empty list.",
    )
    add(
        "13",
        "13 Points Referrals Viral",
        "Ledger running balance",
        "GET /api/ip/points/ledger running balance matches header points",
        "P1",
        "Functional",
        "Candidate, Employer",
        seed,
        "1. Open refer & earn.\n2. Compare ledger sum to user.points.",
        "Chronological sum equals balance. Referral +25 rules: Gmail immediate vs form after SA (REG-C-6/7).",
        "Deterministic seed: do not assert exact 50 if core-fill already spent points.",
    )
    add(
        "18",
        "18 Cross-cutting",
        "Cron auth",
        "Schedule-reminders and export-jobs cron routes are not callable as a normal logged-in user without cron secret",
        "P0",
        "Security",
        "Developer / QA",
        "Read route.js for cron auth before firing",
        "1. GET/POST cron routes as candidate.\n2. Repeat with whatever secret the code expects.",
        "User session alone must not process all reminders/exports. Flag actual mechanism (header vs env) in Actual Result when run.",
        "Do not publish secrets into the sheet.",
    )
    add(
        "08",
        "08 Candidate Pipeline UX",
        "Home widgets",
        "Candidate dashboard cards (ratings, applications, points) load with current schema",
        "P1",
        "Functional",
        "Candidate",
        DEMO["candidate"],
        "1. Open /candidate.\n2. Follow each CTA to profile/browse/offers.",
        "No 500. CTAs hit real routes in ipNav.",
        "Related CAND-D-1 rewritten for Gemini home.",
    )
    add(
        "11",
        "11 Offers Ratings Completions",
        "Offer needs application",
        "POST /api/ip/offers without an application is rejected",
        "P0",
        "Negative",
        "Employer",
        DEMO["employer"] + "; a published internship; candidate who has not applied",
        "1. POST /api/ip/offers with candidateId + internshipId and no applicationId, where no ip_applications row exists.\n2. Repeat with a bogus applicationId.",
        "400: Offer requires an existing application. Bogus applicationId → 404. No ip_offers row is inserted (application_id is NOT NULL).",
        "EP: missing both applicationId and candidate+internship vs candidate+internship with no apply vs valid applicationId. Risk 3×3=9 (orphan offers). Trace: migrations 019/023. Demo: "
        + DEMO["employer"]
        + " / "
        + DEMO["candidate2"]
        + " if +2 has not applied to that posting.",
    )
    add(
        "11",
        "11 Offers Ratings Completions",
        "One offer per application",
        "Second offer on the same application_id is 409",
        "P0",
        "Negative",
        "Employer",
        DEMO["employer"] + " with an application that already has an offer, or create one then POST again",
        "1. POST /api/ip/offers with a valid applicationId (201).\n2. POST the same applicationId again.",
        "Second POST 409: This application already has an offer. Unique ip_offers.application_id.",
        "BVA: 0 offers, 1 offer, 2nd attempt. Do not use Placement Hub college offer upload.",
    )
    add(
        "11",
        "11 Offers Ratings Completions",
        "Accept sets hired",
        "Candidate accept PATCH updates offer accepted and application hired in one step",
        "P0",
        "Functional",
        "Candidate, Employer",
        "Pending offer with application_id (do not use a core offer you cannot restore unless using generate:ip-test-data)",
        "1. Candidate PATCH /api/ip/offers/{id} { status: accepted }.\n2. GET application (employer applicants or DB).",
        "Offer status accepted. Linked application status hired (not left at offered). Employer cannot PATCH accept (CAND-O-4).",
        "Decision table: pending×accepted, pending×declined, already-responded, expired. Migration 029. Test data: generated run id tagged users, not only cores.",
    )
    add(
        "11",
        "11 Offers Ratings Completions",
        "Decline sets declined_offer",
        "Candidate decline PATCH sets application declined_offer",
        "P0",
        "Functional",
        "Candidate",
        "Separate pending offer from accept case",
        "1. PATCH { status: declined }.\n2. Confirm application status declined_offer.",
        "Offer declined. Application declined_offer. Status CHECK allows declined_offer.",
        "EP: decline vs accept vs invalid status hired on offer PATCH.",
    )
    add(
        "11",
        "11 Offers Ratings Completions",
        "Rating/endorsement engagement gate",
        "Rating and endorsement require internshipId and hired or completed application",
        "P0",
        "Negative",
        "Employer, Candidate",
        DEMO["employer"] + "; " + DEMO["candidate"] + " with only applied/shortlisted if possible",
        "1. POST /api/ip/ratings { toUserId, stars: 5 } without internshipId.\n2. POST ratings with internshipId while application is applied.\n3. POST /api/ip/endorsements { candidateId } without internshipId.\n4. After hired/completed, POST rating and endorsement with internshipId.",
        "Steps 1–3: 400 (internshipId required or no hired/completed engagement). Step 4: 201 or 409 if already recorded for that internship.",
        "requireInternshipEngagement ALLOWED={hired, completed}. Endorsement unique employer+candidate+internship.",
    )
    add(
        "11",
        "11 Offers Ratings Completions",
        "Duplicate rating unique",
        "Second rating for same from/to/internship is 409",
        "P1",
        "Negative",
        "Employer",
        "Completed/hired pair already rated once",
        "1. POST rating with internshipId (201 or existing).\n2. POST the same triple again.",
        "409 unique ip_ratings_from_to_internship_key. Not a 500.",
        "BVA: first insert vs duplicate.",
    )
    add(
        "11",
        "11 Offers Ratings Completions",
        "Duplicate endorsement unique",
        "Second endorsement for same employer+candidate+internship is 409",
        "P1",
        "Negative",
        "Employer",
        "Hired/completed application",
        "1. POST /api/ip/endorsements with candidateId + internshipId.\n2. Repeat.",
        "First 201; second 409 You have already endorsed this candidate for this internship.",
        "EP: missing internshipId vs valid vs duplicate.",
    )
    add(
        "12",
        "12 Messages",
        "Dead notification targets",
        "Notifications for deleted internship/offer set resourceUnavailable and clear the link",
        "P1",
        "Functional",
        "Candidate",
        DEMO["candidate"] + "; a notification whose meta.internshipId or offerId no longer exists (or generate one then delete the entity in a throwaway run)",
        "1. GET /api/ip/notifications.\n2. Find or create a notice pointing at a missing internship/offer.\n3. Confirm UI on /candidate/notifications.",
        "annotateNotificationsTargetAvailability: resourceUnavailable true, resourceUnavailableMessage set, link and actionHref null. List routes without a specific id stay clickable.",
        "EP: missing internship vs missing offer vs missing application vs live entity. Risk 2×3=6. Do not delete core showcase internships.",
    )
    add(
        "12",
        "12 Messages",
        "Thread application_id",
        "Message thread may have application_id when the candidate applied; invite-only thread may be null",
        "P1",
        "Functional",
        "Candidate, Employer",
        DEMO["employer"] + " invite vs apply paths",
        "1. Create/open a thread after an application (POST /api/ip/messages/threads with internshipId).\n2. Create/open a thread with no internship (general).\n3. GET thread list.",
        "Applied+internship: application_id set when linkThreadToApplicationIfPresent finds a row. Pre-application invite: application_id null. FK SET NULL if application deleted. candidate_user_id and employer_user_id NOT NULL.",
        "Migration 023. Decision table: internshipId Y/N × application exists Y/N.",
    )
    return out


def style_header(ws, row: int):
    for col, name in enumerate(COLS, 1):
        cell = ws.cell(row=row, column=col, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def write_cases(ws, cases: list[dict], start_row: int):
    for i, c in enumerate(cases):
        r = start_row + i
        values = [
            c["id"],
            c["module"],
            c["feature"],
            c["title"],
            c["priority"],
            c["type"],
            c["roles"],
            c["pre"],
            c["steps"],
            c["expected"],
            c["notes"],
            c["automation"],
            "Not Run",
            "",
            "",
            c.get("legacy") or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 13:
                cell.fill = NOTRUN_FILL
    widths = [16, 28, 22, 52, 10, 12, 28, 36, 44, 44, 44, 22, 12, 22, 16, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    last = start_row + len(cases) - 1
    letter = get_column_letter(13)
    rng = f"{letter}{start_row}:{letter}{max(last, start_row + 200)}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Pass"'], fill=PASS_FILL, font=PASS_FONT))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Fail"'], fill=FAIL_FILL, font=FAIL_FONT))
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"Blocked"'], fill=BLOCK_FILL, font=BLOCK_FONT)
    )
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"Not Run"'], fill=NOTRUN_FILL)
    )


def priority_counts(cases):
    p0 = sum(1 for c in cases if c["priority"] == "P0")
    p1 = sum(1 for c in cases if c["priority"] == "P1")
    p2 = sum(1 for c in cases if c["priority"] not in ("P0", "P1"))
    return len(cases), p0, p1, p2


MODULE_ORDER = [
    "01 Public & Landing",
    "02 Auth & Access",
    "03 Registration",
    "04 Permissions & Role",
    "05 Account & Security",
    "06 Candidate Profile",
    "07 Browse Save Apply",
    "08 Candidate Pipeline UX",
    "09 Employer Postings Pipeline",
    "10 Employer Candidate Search",
    "11 Offers Ratings Completions",
    "12 Messages",
    "13 Points Referrals Viral",
    "14 SuperAdmin Ops",
    "15 Feature Ideas",
    "16 List Filters Sort Presets",
    "17 Files & Uploads",
    "18 Cross-cutting",
]

COVERAGE_ROWS = [
    ("Sign-in on `/` + captcha + 2FA", "Y", "Y", "Y (/superadmin/login)", "Y"),
    ("Register (Gmail candidate / employer domain or request)", "Y", "Y", "approve", "Y"),
    ("Full-page candidate profile tabs", "Y", "view /[id]", "—", "—"),
    ("Browse / save / apply (5 pts)", "Y", "—", "—", "—"),
    ("Postings + applicant pipeline + bulk/export", "—", "Y", "oversee", "—"),
    ("List last-used prefs + presets (max 5)", "Y", "Y", "API 403", "—"),
    ("Offers / ratings / completions + application_id integrity", "Y", "Y", "—", "—"),
    ("Messages + attachments", "Y", "Y", "oversight", "—"),
    ("Notifications UI", "Y", "Y", "API only / no nav page", "—"),
    ("Points / referral /r/{code}", "Y", "Y", "—", "landing"),
    ("Points convert", "NO route", "NO route", "—", "—"),
    ("Viral / LinkedIn promos", "—", "Y", "Y", "—"),
    ("Feature ideas", "Y", "Y", "moderate", "—"),
    ("Account 2FA / sessions / contact change", "Y", "Y", "Y", "—"),
]


def main():
    payload = json.loads(DUMP.read_text(encoding="utf-8"))
    b_cases = payload["cases"]
    if len(b_cases) != 179:
        raise SystemExit(f"Expected 179 Reference B cases, got {len(b_cases)}")

    seq = defaultdict(lambda: 1)
    modules = {m: [] for m in MODULE_ORDER}
    recon_rows = []

    seen = set()
    for row in b_cases:
        lid = row["id"]
        if lid in seen:
            raise SystemExit("duplicate " + lid)
        seen.add(lid)
        rec = rewrite_from_b(row, seq)
        modules[rec["module"]].append(rec)
        recon_rows.append(
            {
                "legacy_id": lid,
                "old_sheet": row["sheet"],
                "classification": rec["recon"],
                "new_tc_id": rec["id"],
                "new_module": rec["module"],
                "drop_reason": None,
            }
        )

    for rec in extra_cases(seq):
        modules.setdefault(rec["module"], []).append(rec)
        if rec["module"] not in MODULE_ORDER:
            MODULE_ORDER.append(rec["module"])

    # Dropped count
    dropped = [r for r in recon_rows if r["classification"].startswith("Dropped")]
    if dropped:
        raise SystemExit("Unexpected drops")

    wb = Workbook()
    idx = wb.active
    idx.title = "Index"
    idx["A1"] = "InternSafar — Test Case Index"
    idx["A1"].font = INDEX_TITLE_FONT
    idx["A2"] = (
        "Product: InternSafar (workspace sibling internship-portal). "
        "Column schema matches Reference A (TC ID … Executed At) plus Legacy ID for eyeballing Reference B. "
        "Reference A's test CONTENT was not copied. Status is Not Run for every row. "
        "Reference B (~179 executable IDs in the xlsx, not 212) was a one-time content source and is not linked. "
        "Integrity cases (offer↔application, rating/endorsement engagement, dead notification targets, thread application_id) "
        "were added 25 Aug 2026 for current InternSafar APIs. Every Status is Not Run."
    )
    idx["A2"].alignment = Alignment(wrap_text=True)
    idx.merge_cells("A2:G2")
    idx.row_dimensions[2].height = 56
    headers = ["#", "Module / Tab (click to open)", "Case Count", "P0", "P1", "P2+", "Link"]
    for col, h in enumerate(headers, 1):
        cell = idx.cell(row=4, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    idx["A5"] = 0
    idx["B5"] = "Coverage Matrix"
    idx["C5"] = 0
    idx["G5"] = "Open >"
    idx["G5"].font = LINK_FONT
    idx["G5"].hyperlink = "#'Coverage Matrix'!A1"

    row = 6
    total = p0t = p1t = p2t = 0
    n = 1
    for mod in MODULE_ORDER:
        cases = modules.get(mod) or []
        if not cases:
            continue
        count, p0, p1, p2 = priority_counts(cases)
        total += count
        p0t += p0
        p1t += p1
        p2t += p2
        idx.cell(row=row, column=1, value=n)
        idx.cell(row=row, column=2, value=mod)
        idx.cell(row=row, column=3, value=count)
        idx.cell(row=row, column=4, value=p0)
        idx.cell(row=row, column=5, value=p1)
        idx.cell(row=row, column=6, value=p2)
        link = idx.cell(row=row, column=7, value="Open >")
        link.font = LINK_FONT
        link.hyperlink = f"#'{mod[:31]}'!A1"
        n += 1
        row += 1

    idx.cell(row=row, column=2, value="TOTAL")
    idx.cell(row=row, column=3, value=total)
    idx.cell(row=row, column=4, value=p0t)
    idx.cell(row=row, column=5, value=p1t)
    idx.cell(row=row, column=6, value=p2t)
    row += 2
    idx.cell(row=row, column=1, value="Demo logins")
    idx.cell(
        row=row,
        column=2,
        value=f"Candidate: {DEMO['candidate']} | Employer: {DEMO['employer']} | SA: {DEMO['superadmin']}",
    )
    row += 1
    idx.cell(row=row, column=1, value="Priority")
    idx.cell(row=row, column=2, value="P0 risk score ≥6 or auth/economy/security | P1 high workflows | P2+ polish")
    row += 1
    idx.cell(row=row, column=1, value="Extra column")
    idx.cell(
        row=row,
        column=2,
        value="Legacy ID is an annotation requested for mapping; it is not a Reference B schema leftover (Doc Status/Phase/Dev comment omitted).",
    )
    idx.column_dimensions["A"].width = 14
    idx.column_dimensions["B"].width = 38
    idx.column_dimensions["C"].width = 12
    idx.column_dimensions["G"].width = 10

    cov = wb.create_sheet("Coverage Matrix")
    cov["A1"] = "<< Back to Index"
    cov["A1"].font = LINK_FONT
    cov["A1"].hyperlink = "#Index!A1"
    cov["A2"] = "Role × capability (InternSafar current sibling)"
    for col, h in enumerate(["Capability", "Candidate", "Employer", "SuperAdmin", "Public"], 1):
        cell = cov.cell(row=3, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, r in enumerate(COVERAGE_ROWS):
        for col, val in enumerate(r, 1):
            cov.cell(row=4 + i, column=col, value=val)
    for i, w in enumerate([48, 22, 22, 28, 12], 1):
        cov.column_dimensions[get_column_letter(i)].width = w

    for mod in MODULE_ORDER:
        cases = modules.get(mod) or []
        if not cases:
            continue
        ws = wb.create_sheet(mod[:31])
        ws["A1"] = "<< Back to Index"
        ws["A1"].font = LINK_FONT
        ws["A1"].hyperlink = "#Index!A1"
        style_header(ws, 2)
        write_cases(ws, cases, 3)
        ws.auto_filter.ref = f"A2:{get_column_letter(len(COLS))}{2 + len(cases)}"
        ws.freeze_panes = "A3"
        ws.row_dimensions[2].height = 22
        ws.sheet_properties.pageSetUpPr.fitToPage = True

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT} cases={total} P0={p0t} P1={p1t} P2+={p2t}")


if __name__ == "__main__":
    main()
