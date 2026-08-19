# -*- coding: utf-8 -*-
"""Apply automated QA results to test-cases/Internship_Portal_Test_Checklist.xlsx."""
from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
CHECKLIST = ROOT / "test-cases" / "Internship_Portal_Test_Checklist.xlsx"
SKIP_SHEETS = frozenset({"Index", "Coverage Matrix", "Coverage by Type"})

PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
PASS_FONT = Font(bold=True, color="006100")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
FAIL_FONT = Font(bold=True, color="9C0006")
BLOCK_FILL = PatternFill("solid", fgColor="D9D9D9")
BLOCK_FONT = Font(bold=True, color="595959")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")

STYLE = {
    "Pass": (PASS_FILL, PASS_FONT),
    "Fail": (FAIL_FILL, FAIL_FONT),
    "Blocked": (BLOCK_FILL, BLOCK_FONT),
    "In Progress": (PatternFill("solid", fgColor="FFF2CC"), Font(bold=True, color="9C6500")),
}


def find_header_row(ws) -> int:
    for r in range(1, 8):
        if ws.cell(r, 1).value == "ID" and ws.cell(r, 4).value == "Issue Summary":
            return r
    return 3


def apply_status_style(cell):
    val = (cell.value or "").strip() if isinstance(cell.value, str) else cell.value
    if val in STYLE:
        fill, font = STYLE[val]
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)


def ensure_conditional_formatting(ws, header_row: int, status_col: int, max_row: int):
    if max_row < header_row + 1:
        return
    letter = get_column_letter(status_col)
    rng = f"{letter}{header_row + 1}:{letter}{max(max_row, header_row + 200)}"
    try:
        for sqref in list(ws.conditional_formatting._cf_rules.keys()):
            if letter in str(sqref):
                del ws.conditional_formatting[sqref]
    except Exception:
        pass
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"Pass"'], fill=PASS_FILL, font=PASS_FONT)
    )
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"Fail"'], fill=FAIL_FILL, font=FAIL_FONT)
    )
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"Blocked"'], fill=BLOCK_FILL, font=BLOCK_FONT)
    )


def style_sheet(ws):
    hr = find_header_row(ws)
    header = [ws.cell(hr, c).value for c in range(1, ws.max_column + 1)]
    if "Test Status" not in header:
        return
    sc = header.index("Test Status") + 1
    hcell = ws.cell(hr, sc)
    hcell.fill = HEADER_FILL
    hcell.font = HEADER_FONT
    max_r = ws.max_row or hr
    for r in range(hr + 1, max_r + 1):
        apply_status_style(ws.cell(r, sc))
    ensure_conditional_formatting(ws, hr, sc, max_r)


def normalize_cases(raw: dict) -> dict[str, dict]:
    out: dict[str, dict] = {}
    for key, val in raw.items():
        if not isinstance(val, dict):
            continue
        status = val.get("status") or val.get("testStatus") or ""
        actual = val.get("actual") or val.get("comments") or val.get("notes") or ""
        out[str(key).strip()] = {"status": status, "actual": actual}
    return out


def update_workbook(wb, cases: dict[str, dict], executed_at: str) -> list[str]:
    updated: list[str] = []
    date_str = executed_at[:10] if executed_at else datetime.now(timezone.utc).isoformat()[:10]

    for name in wb.sheetnames:
        if name in SKIP_SHEETS:
            continue
        ws = wb[name]
        hr = find_header_row(ws)
        header = [ws.cell(hr, c).value for c in range(1, ws.max_column + 1)]
        col = {name: i + 1 for i, name in enumerate(header) if name}
        if "ID" not in col or "Test Status" not in col:
            continue

        for row in range(hr + 1, ws.max_row + 1):
            case_id = ws.cell(row, col["ID"]).value
            if not case_id or str(case_id).strip() not in cases:
                continue
            case_id = str(case_id).strip()
            r = cases[case_id]
            status = r.get("status") or "Fail"
            ws.cell(row, col["Test Status"], value=status)
            apply_status_style(ws.cell(row, col["Test Status"]))
            if "Date Verified" in col and status in ("Pass", "Fail", "Blocked"):
                ws.cell(row, col["Date Verified"], value=date_str)
            if "Comments / Notes" in col and r.get("actual"):
                prev = ws.cell(row, col["Comments / Notes"]).value
                note = str(r["actual"])
                if prev and str(prev).strip() and "[auto]" not in str(prev):
                    note = f"[auto] {note}\n\n{prev}"
                else:
                    note = f"[auto] {note}"
                ws.cell(row, col["Comments / Notes"], value=note)
                ws.cell(row, col["Comments / Notes"]).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )
            updated.append(case_id)

    for name in wb.sheetnames:
        if name not in SKIP_SHEETS:
            style_sheet(wb[name])
    return updated


def load_payload(path: Path) -> tuple[dict, str]:
    raw = path.read_text(encoding="utf-8-sig")
    start = raw.find("{")
    data = json.loads(raw[start:] if start >= 0 else raw)
    executed_at = data.get("executedAt") or datetime.now(timezone.utc).isoformat()
    if "cases" in data:
        return normalize_cases(data["cases"]), executed_at
    if "batches" in data:
        merged: dict[str, dict] = {}
        for batch in data["batches"]:
            merged.update(normalize_cases(batch.get("cases") or {}))
        return merged, data.get("executedAt") or executed_at
    return normalize_cases(data), executed_at


def main():
    if not CHECKLIST.exists():
        print(f"Missing checklist: {CHECKLIST}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(CHECKLIST)
    updated: list[str] = []

    if len(sys.argv) > 1:
        payload_path = Path(sys.argv[1])
        if not payload_path.is_absolute():
            payload_path = ROOT / payload_path
        cases, executed_at = load_payload(payload_path)
        updated = update_workbook(wb, cases, executed_at)
        print("updated", len(updated), "cases")
        print("ids", updated[:20], "..." if len(updated) > 20 else "")

    wb.save(CHECKLIST)
    print("saved", CHECKLIST)


if __name__ == "__main__":
    main()
